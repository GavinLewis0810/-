import hashlib
import json
import logging
from datetime import datetime

from fastapi import APIRouter, Depends, HTTPException, status, Request
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func
from sqlalchemy.orm import selectinload
from typing import List, Optional
from decimal import Decimal

from app.database import get_db

logger = logging.getLogger(__name__)
from app.models.reimbursement import Reimbursement, ReimbursementStatus
from app.models.invoice import Invoice, InvoiceStatus, ParsingDiff
from app.models.user import User
from app.models.audit_log import AuditLog
from app.models.notification import Notification
from app.models.project import Project
from app.models.application import Application, ApplicationStatus
from app.models.bank_card import BankCard
from app.models.transaction import Transaction
from app.schemas.reimbursement import (
    ReimbursementCreate,
    ReimbursementResponse,
    ReimbursementReview,
    CategorySuggestionRequest,
    CategorySuggestionResponse,
    VoucherReviewRequest,
    VoucherReviewResponse,
)
from app.services.reimbursement_service import delete_reimbursement_logic
from app.services.audit_service import log_audit_no_commit, get_client_info
from app.services.ws_manager import push_notification, push_notification_to_admins
from app.services.carbon_config import SPEND_TO_REASON_MAP
from app.models.reason_category import ReasonCategory
from app.dependencies import get_current_user

router = APIRouter()


def _serialize_invoice_field(value):
    if value is None:
        return None
    if hasattr(value, "isoformat"):
        return value.isoformat()
    if isinstance(value, Decimal):
        return str(value)
    return str(value)


def _set_invoice_field(invoice: Invoice, field_name: str, raw_value: str):
    if field_name == "issue_date":
        invoice.issue_date = datetime.strptime(raw_value, "%Y-%m-%d").date()
        return
    if field_name in {"total_with_tax", "amount", "tax_amount"}:
        setattr(invoice, field_name, Decimal(str(raw_value)))
        return
    setattr(invoice, field_name, raw_value)


def _rebuild_invoice_hash(invoice: Invoice):
    raw = json.dumps(
        {
            "invoice_number": invoice.invoice_number,
            "issue_date": str(invoice.issue_date) if invoice.issue_date else None,
            "total_with_tax": str(invoice.total_with_tax) if invoice.total_with_tax else None,
            "amount": str(invoice.amount) if invoice.amount else None,
            "tax_amount": str(invoice.tax_amount) if invoice.tax_amount else None,
            "buyer_name": invoice.buyer_name,
            "buyer_tax_id": invoice.buyer_tax_id,
            "seller_name": invoice.seller_name,
            "seller_tax_id": invoice.seller_tax_id,
            "items": json.dumps(invoice.items, ensure_ascii=False) if invoice.items else None,
            "owner_id": invoice.owner_id,
            "status": invoice.status.value if invoice.status else None,
        },
        sort_keys=True,
        ensure_ascii=False,
    )
    invoice.invoice_hash = hashlib.sha256(raw.encode("utf-8")).hexdigest()


def _invoice_needs_voucher_review(invoice: Invoice) -> bool:
    if invoice.status == InvoiceStatus.PENDING_VOUCHER_REVIEW:
        return True
    return invoice.confirmation_mode in {"USER_SELECTION", "ADMIN_CORRECTION"}


def _invoice_reviewed(invoice: Invoice) -> bool:
    trace = invoice.decision_trace or {}
    review = trace.get("voucher_review") or {}
    return bool(review.get("reviewed"))


@router.post("/category-suggestion", response_model=CategorySuggestionResponse)
async def suggest_category(
    data: CategorySuggestionRequest,
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """根据选中发票的消费类别，智能建议报销事由类别"""
    # 1. 如果有申请单，直接继承（现有逻辑，这里返回提示）
    if data.application_id:
        app = await db.get(Application, data.application_id)
        if app and app.reason_category_id:
            rc = await db.get(ReasonCategory, app.reason_category_id)
            return CategorySuggestionResponse(
                mode="application_override",
                suggested_category_id=app.reason_category_id,
                suggested_category_name=rc.name if rc else None,
                confidence=1.0,
                breakdown=[],
                hint=f"已关联申请单，事由类别从申请单继承：{rc.name if rc else '未知'}",
            )

    # 2. 查所有选中发票的 spend_category
    invoices_result = await db.execute(
        select(Invoice.spend_category).where(Invoice.id.in_(data.invoice_ids))
    )
    categories = [row[0] for row in invoices_result.fetchall() if row[0]]

    if not categories:
        return CategorySuggestionResponse(
            mode="suggestion",
            hint="所选发票尚未完成智能分类，请手动选择事由类别",
        )

    # 3. 每张发票映射到 reason_category
    mapped: dict[str, int] = {}  # reason_category_name → count
    breakdown = []
    for cat in categories:
        reason_name = SPEND_TO_REASON_MAP.get(cat, "其他费用")
        mapped[reason_name] = mapped.get(reason_name, 0) + 1
        breakdown.append({
            "spend_category": cat,
            "mapped_category": reason_name,
        })

    best = max(mapped, key=mapped.get)
    confidence = mapped[best] / len(categories)

    # 4. 查 reason_category 的实际 ID
    rc_result = await db.execute(
        select(ReasonCategory).where(ReasonCategory.name == best, ReasonCategory.is_active == True)
    )
    rc = rc_result.scalar_one_or_none()

    detail_parts = [f"{b['spend_category']}→{b['mapped_category']}" for b in breakdown]
    hint = f"根据 {len(categories)} 张发票内容建议：{best}（{' / '.join(detail_parts)}）"

    return CategorySuggestionResponse(
        mode="suggestion",
        suggested_category_id=rc.id if rc else None,
        suggested_category_name=best,
        confidence=round(confidence, 2),
        breakdown=breakdown,
        hint=hint,
    )


@router.post("", response_model=ReimbursementResponse)
async def create_reimbursement(
    data: ReimbursementCreate,
    request: Request,
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    query = select(Invoice).where(Invoice.id.in_(data.invoice_ids))
    result = await db.execute(query)
    invoices = result.scalars().all()

    if not invoices:
        raise HTTPException(status_code=404, detail="未找到指定的发票")

    total = sum([inv.total_with_tax for inv in invoices if inv.total_with_tax])
    total_carbon = sum([float(inv.carbon_kg or 0) for inv in invoices])

    # 预算超标检查
    if data.project_code:
        project = await db.scalar(select(Project).where(Project.project_code == data.project_code))
        if project and project.budget > 0:
            used = await db.scalar(
                select(func.coalesce(func.sum(Reimbursement.total_amount), 0)).where(
                    Reimbursement.project_code == data.project_code,
                    Reimbursement.status.in_([ReimbursementStatus.SUBMITTED, ReimbursementStatus.APPROVED, ReimbursementStatus.COMPLETED]),
                )
            ) or 0
            if Decimal(str(used)) + total > project.budget:
                remaining = project.budget - Decimal(str(used))
                raise HTTPException(
                    status_code=400,
                    detail=f"报销金额 ¥{float(total):.2f} 超出项目「{project.project_name}」剩余预算 ¥{float(remaining):.2f}（总预算 ¥{float(project.budget):.2f}，已使用 ¥{float(used):.2f}）",
                )

    # 关联申请单校验：归属 + 状态 + 剩余额度
    reason_category_id = data.reason_category_id
    if data.application_id:
        app = await db.get(Application, data.application_id)
        if not app:
            raise HTTPException(status_code=400, detail="申请单不存在")
        if app.user_id != current_user["id"]:
            raise HTTPException(status_code=400, detail="申请单不属于您")
        if app.status != ApplicationStatus.APPROVED:
            raise HTTPException(status_code=400, detail="只能关联状态为「已通过」的事前申请单")

        # 自动继承申请单的事由类别
        if app.reason_category_id and not reason_category_id:
            reason_category_id = app.reason_category_id

        # 计算该申请单已被占用的额度（已通过+已打款的报销单）
        app_used = await db.scalar(
            select(func.coalesce(func.sum(Reimbursement.total_amount), 0)).where(
                Reimbursement.application_id == data.application_id,
                Reimbursement.status.in_([ReimbursementStatus.APPROVED, ReimbursementStatus.COMPLETED]),
            )
        ) or 0
        # 自动继承申请单的项目编号
        if app.project_code and not data.project_code:
            data.project_code = app.project_code

        app_remaining = app.estimated_amount - Decimal(str(app_used))
        if Decimal(str(total)) > app_remaining:
            raise HTTPException(
                status_code=400,
                detail=f"报销金额 ¥{float(total):.2f} 超出申请单剩余额度 ¥{float(app_remaining):.2f}"
                      f"（申请总额 ¥{float(app.estimated_amount):.2f}，已使用 ¥{float(app_used):.2f}）",
            )
    # 关联借款申请校验
    if data.borrowing_id:
        from app.models.borrowing import Borrowing, BorrowingStatus
        borrowing = await db.get(Borrowing, data.borrowing_id)
        if not borrowing:
            raise HTTPException(status_code=400, detail="关联的借款申请不存在")
        if borrowing.user_id != current_user["id"]:
            raise HTTPException(status_code=400, detail="借款申请不属于您")
        if borrowing.status != BorrowingStatus.APPROVED.value:
            raise HTTPException(status_code=400, detail="只能关联状态为「已批准」的借款申请")
        if borrowing.reimbursement_id is not None:
            raise HTTPException(status_code=400, detail="该借款已被其他报销单冲销")

    reimb = Reimbursement(
        title=data.title,
        project_code=data.project_code,
        total_amount=total,
        carbon_kg=round(total_carbon, 4) if total_carbon > 0 else None,
        submitter=current_user["username"],
        submitter_id=current_user["id"],
        bank_card_id=data.bank_card_id,
        application_id=data.application_id,
        borrowing_id=data.borrowing_id,
        reason_category_id=reason_category_id,
        status=ReimbursementStatus.SUBMITTED
    )
    db.add(reimb)
    await db.flush()

    for inv in invoices:
        # 防止拿别人的发票报销
        if inv.owner_id and inv.owner_id != current_user["id"]:
            raise HTTPException(status_code=403, detail=f"发票 {inv.invoice_number or inv.id} 不属于您，无法报销")
        # 确认流门禁：待重审必须先管理员复核；已确认和待随单审核可进报销单
        if inv.status == InvoiceStatus.PENDING_RECHECK:
            raise HTTPException(status_code=400, detail=f"发票 {inv.invoice_number or inv.id} 处于待重审，暂不可报销")
        if inv.status not in [InvoiceStatus.CONFIRMED, InvoiceStatus.PENDING_VOUCHER_REVIEW]:
            raise HTTPException(status_code=400, detail=f"发票 {inv.invoice_number or inv.id} 状态不支持报销")
        # 防止同一张发票重复打包
        if inv.reimbursement_id is not None:
            raise HTTPException(status_code=400, detail=f"发票 {inv.invoice_number or inv.id} 已被其他报销单占用")
        # 按发票号码去重：同一号码不能同时被多人使用
        if inv.invoice_number:
            dup = await db.scalar(
                select(func.count(Invoice.id)).where(
                    Invoice.invoice_number == inv.invoice_number,
                    Invoice.id != inv.id,
                    Invoice.reimbursement_id.isnot(None),
                )
            )
            if dup and dup > 0:
                raise HTTPException(status_code=400, detail=f"发票号码 {inv.invoice_number} 已被他人提交报销，无法重复使用")
        inv.reimbursement_id = reimb.id

    await push_notification_to_admins(
        db,
        title="新报销单待审批",
        message=f"员工 {current_user.get('full_name') or current_user.get('username')} 提交了报销单「{reimb.title}」，金额 ¥{float(reimb.total_amount or 0):.2f}，请及时审批。",
        entity_type="reimbursement",
        entity_id=reimb.id,
    )

    client_info = get_client_info(request)
    await log_audit_no_commit(
        db=db,
        entity_type="reimbursement",
        entity_id=reimb.id,
        action="submit",
        new_value={"status": ReimbursementStatus.SUBMITTED.value, "amount": float(reimb.total_amount or 0)},
        ip_address=client_info.get("ip_address"),
        user_agent=client_info.get("user_agent"),
    )

    await db.commit()

    fetch_query = select(Reimbursement).options(selectinload(Reimbursement.invoices)).where(
        Reimbursement.id == reimb.id)
    fetch_result = await db.execute(fetch_query)
    return fetch_result.scalar_one()


@router.get("", response_model=List[ReimbursementResponse])
async def get_reimbursements(
    skip: int = 0,
    limit: int = 100,
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """获取报销单列表。管理员看全部，员工只看自己的。"""
    query = select(Reimbursement).options(selectinload(Reimbursement.invoices)).order_by(
        Reimbursement.created_at.desc())

    # 数据隔离：员工只能看自己提交的报销单（通过外键）
    if current_user["role"] != "admin":
        query = query.where(Reimbursement.submitter_id == current_user["id"])

    query = query.offset(skip).limit(limit)
    result = await db.execute(query)
    return result.scalars().all()


@router.get("/{reimb_id}", response_model=ReimbursementResponse)
async def get_reimbursement_detail(
    reimb_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """获取单个报销单详情（含关联发票完整信息）。"""
    from app.models.borrowing import Borrowing
    query = select(Reimbursement).options(
        selectinload(Reimbursement.invoices),
        selectinload(Reimbursement.bank_card),
        selectinload(Reimbursement.borrowing).selectinload(Borrowing.application),
    ).where(Reimbursement.id == reimb_id)

    # 数据隔离
    if current_user["role"] != "admin":
        query = query.where(Reimbursement.submitter_id == current_user["id"])

    result = await db.execute(query)
    reimb = result.scalar_one_or_none()
    if not reimb:
        raise HTTPException(status_code=404, detail="报销单不存在")

    # 计算 bank_card_info 供前端打印凭证使用
    if reimb.bank_card:
        masked = reimb.bank_card.card_number[-4:] if len(reimb.bank_card.card_number) >= 4 else reimb.bank_card.card_number
        reimb.bank_card_info = f"{reimb.bank_card.bank_name} (尾号{masked})"
    elif reimb.payment_bank:
        reimb.bank_card_info = reimb.payment_bank
    else:
        reimb.bank_card_info = None

    # 财务总监电子签名（查 admin 用户）
    reviewer_sig = None
    from app.models.user import User
    admin_query = select(User).where(User.role == "admin").limit(1)
    admin_result = await db.execute(admin_query)
    admin_user = admin_result.scalar_one_or_none()
    if admin_user and admin_user.signature:
        reviewer_sig = admin_user.signature

    # 手动构造返回 dict，确保 reviewer_signature 不会被 Pydantic 过滤掉
    from app.schemas.reimbursement import ReimbursementResponse
    resp = ReimbursementResponse.model_validate(reimb)
    resp_dict = resp.model_dump()
    resp_dict["reviewer_signature"] = reviewer_sig
    # 借款冲销信息
    if reimb.borrowing:
        resp_dict["borrowing_info"] = {
            "id": reimb.borrowing.id,
            "title": reimb.borrowing.title,
            "estimated_amount": float(reimb.borrowing.estimated_amount),
            "repaid_amount": float(reimb.borrowing.repaid_amount) if reimb.borrowing.repaid_amount else None,
            "status": reimb.borrowing.status,
        }
    else:
        resp_dict["borrowing_info"] = None
    return resp_dict


@router.get("/export/excel")
async def export_reimbursements_excel(
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """导出报销单为 Excel 格式。"""
    from io import BytesIO
    from urllib.parse import quote
    from fastapi.responses import StreamingResponse

    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    except ImportError:
        raise HTTPException(status_code=500, detail="Excel 导出需要 openpyxl 库")

    query = select(Reimbursement).options(
        selectinload(Reimbursement.invoices),
        selectinload(Reimbursement.submitter_user),
    ).order_by(Reimbursement.created_at.desc())

    if current_user["role"] != "admin":
        query = query.where(Reimbursement.submitter_id == current_user["id"])

    result = await db.execute(query)
    reimbursements = result.scalars().all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "报销单台账"

    headers = [
        "报销单号", "报销事由", "项目编号", "报销总金额", "提交人",
        "审批人", "审批意见", "驳回理由", "状态", "AI风险评级",
        "关联发票数", "关联发票号码", "提交时间", "更新时间",
    ]

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="E42313", end_color="E42313", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for row, r in enumerate(reimbursements, 2):
        invoice_nums = ", ".join(
            [inv.invoice_number for inv in (r.invoices or []) if inv.invoice_number]
        ) or "-"

        data = [
            r.id,
            r.title,
            r.project_code or "-",
            float(r.total_amount or 0),
            r.submitter_name or r.submitter or "-",
            r.reviewer or "-",
            r.review_note or "-",
            r.reject_reason or "-",
            r.status.value if r.status else "-",
            r.ai_risk_level or "-",
            len(r.invoices or []),
            invoice_nums,
            r.created_at.strftime("%Y-%m-%d %H:%M") if r.created_at else "-",
            r.updated_at.strftime("%Y-%m-%d %H:%M") if r.updated_at else "-",
        ]

        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

    # 自适应列宽
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

    # 合计行
    total_row = len(reimbursements) + 2
    ws.cell(row=total_row, column=1, value="合计").font = Font(bold=True)
    ws.cell(row=total_row, column=4, value=sum(float(r.total_amount or 0) for r in reimbursements))
    ws.cell(row=total_row, column=4).font = Font(bold=True, color="E42313")

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = quote("报销单台账.xlsx")
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"},
    )
async def review_reimbursement(reimb_id: int, data: ReimbursementReview, db: AsyncSession = Depends(get_db)):
    query = select(Reimbursement).options(selectinload(Reimbursement.invoices)).where(Reimbursement.id == reimb_id)
    result = await db.execute(query)
    reimb = result.scalar_one_or_none()

    if not reimb:
        raise HTTPException(status_code=404, detail="报销单不存在")

    if data.action == "APPROVE":
        reimb.status = ReimbursementStatus.APPROVED
    elif data.action == "REJECT":
        reimb.status = ReimbursementStatus.REJECTED
        reimb.reject_reason = data.reject_reason
        for inv in reimb.invoices:
            inv.reimbursement_id = None
            inv.status = InvoiceStatus.CONFIRMED

    await db.commit()
    await db.refresh(reimb)
    return reimb


@router.delete("/{reimb_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_reimbursement(
    reimb_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    success = await delete_reimbursement_logic(
        reimb_id, db,
        deleted_by_username=current_user.get("full_name") or current_user.get("username", "管理员"),
    )
    if not success:
        raise HTTPException(status_code=404, detail="报销单不存在")
    return None


@router.post("/{reimb_id}/ai-check")
async def ai_check_reimbursement(
    reimb_id: int,
    db: AsyncSession = Depends(get_db)
):
    """
    AI 智能合规审查
    """
    import json
    import logging
    from app.services.prompts import build_ai_check_prompt
    from app.services.llm_service import get_llm_service

    logger = logging.getLogger(__name__)

    # 1. 查报销单 + 预加载关联发票
    query = select(Reimbursement).options(
        selectinload(Reimbursement.invoices)
    ).where(Reimbursement.id == reimb_id)
    result = await db.execute(query)
    reimb = result.scalar_one_or_none()

    if not reimb:
        raise HTTPException(status_code=404, detail="报销单不存在")
    if not reimb.invoices:
        raise HTTPException(status_code=400, detail="该报销单没有关联发票，无法审查")

    # 2. 提取发票数据
    invoices_data = []
    for inv in reimb.invoices:
        invoices_data.append({
            "invoice_number": inv.invoice_number or "无号码",
            "seller_name": inv.seller_name or "",
            "total_with_tax": str(inv.total_with_tax) if inv.total_with_tax else "0.00",
            "items": inv.items or [],
        })

    # 3. 检查预算
    budget_info = ""
    if reimb.project_code:
        project = await db.scalar(select(Project).where(Project.project_code == reimb.project_code))
        if project and project.budget > 0:
            used = await db.scalar(
                select(func.coalesce(func.sum(Reimbursement.total_amount), 0)).where(
                    Reimbursement.project_code == reimb.project_code,
                    Reimbursement.status.in_([ReimbursementStatus.SUBMITTED, ReimbursementStatus.APPROVED, ReimbursementStatus.COMPLETED]),
                )
            ) or 0
            remaining = project.budget - Decimal(str(used))
            budget_info = f"\n## 项目预算信息\n- 项目：{project.project_name}（{project.project_code}）\n- 预算总额：¥{float(project.budget):.2f}\n- 已使用：¥{float(used):.2f}\n- 剩余：¥{float(remaining):.2f}\n- {'⚠️ 注意：本次报销将超出项目预算！' if remaining < reimb.total_amount else '预算充足'}"

    # 4. 构造 Prompt
    prompt = build_ai_check_prompt(
        reimb_title=reimb.title,
        reimb_amount=str(reimb.total_amount),
        invoices=invoices_data,
        budget_info=budget_info,
    )

    # 4. 调用 LLM（使用现有的 QwenProvider）
    try:
        llm_service = get_llm_service()
        provider = llm_service.active_provider
        if not provider or not provider.is_configured():
            raise HTTPException(status_code=500, detail="LLM 服务未配置")

        llm_response = provider.chat_completion(
            system_prompt="你是一位严格的财务审计专家，你的职责是找出报销单中的一切可疑之处。默认持怀疑态度，只有证据充分时才判定合规。必须严格按照 JSON 格式输出。",
            user_prompt=prompt
        )
    except Exception as exc:
        logger.error(f"LLM 调用失败：{exc}")
        raise HTTPException(status_code=500, detail=f"AI 审查服务调用失败：{str(exc)}")

    # 5. 解析结果
    cleaned = llm_response.strip()
    if cleaned.startswith("```"):
        lines = cleaned.splitlines()
        cleaned = "\n".join(lines[1:-1]) if len(lines) > 2 else lines[0]

    try:
        result_dict = json.loads(cleaned)
    except json.JSONDecodeError:
        logger.warning(f"AI 返回非 JSON 格式，原文：{llm_response}")
        result_dict = {
            "compliance_status": "未知",
            "risk_level": "未知",
            "reason": "AI 返回格式异常，请查看备注",
            "remarks": llm_response,
            "details": [],
        }

    # 🚀🚀🚀 核心逻辑：把 AI 的完整审查报告永久刻在数据库里！
    reimb.ai_risk_level = result_dict.get("risk_level", "未知")
    
    # 🛑 核心修复：绝对不能只存一段纯文本！必须把整个字典转化为 JSON 字符串，喂给 ai_reason！
    reimb.ai_reason = json.dumps(result_dict, ensure_ascii=False)
    
    # 兼容原有的字段（如果有的话）
    if hasattr(reimb, 'ai_review_detail'):
        reimb.ai_review_detail = result_dict 

    await db.commit()  # 提交到 PostgreSQL 数据库保存！

    # 动态规则引擎：AI 审查完成后自动匹配规则
    from app.services.rule_engine import match_rules
    from app.models.notification import Notification

    action = await match_rules("reimbursement", {
        "total_amount": float(reimb.total_amount or 0),
        "ai_risk_level": result_dict.get("risk_level", ""),
        "compliance_status": result_dict.get("compliance_status", ""),
    }, db)

    if action == "AUTO_APPROVE":
        # 随单审核门禁：有发票需人工复核则阻止自动通过
        needs_voucher_review = any(
            _invoice_needs_voucher_review(inv) for inv in reimb.invoices
        )
        if needs_voucher_review:
            result_dict["auto_approve_blocked"] = True
            result_dict["auto_approve_block_reason"] = (
                "该报销单满足自动审批条件，但因包含需随单审核的发票字段"
                "（用户选择了非默认字段值或进行了手动修正），已转为人工审批"
            )
            reimb.ai_reason = json.dumps(result_dict, ensure_ascii=False)
            await db.commit()
            await push_notification_to_admins(
                db,
                title="报销单需人工审批",
                message=f"「{reimb.title}」¥{float(reimb.total_amount or 0):.2f} AI审查合规，但因含需随单审核的发票，已阻止自动通过，请前往审批。",
                entity_type="reimbursement", entity_id=reimb.id,
            )
            return result_dict

        # 无随单审核需求，正常自动通过
        reimb.status = ReimbursementStatus.APPROVED
        reimb.reviewer = "AI规则引擎"
        for inv in reimb.invoices:
            inv.status = InvoiceStatus.REIMBURSED
        if reimb.submitter_id:
            await push_notification(
                db, reimb.submitter_id,
                title="报销单自动审批通过",
                message=f"您的报销单「{reimb.title}」¥{float(reimb.total_amount or 0):.2f} 已由AI规则引擎自动审批通过，进入待打款队列。",
                entity_type="reimbursement", entity_id=reimb.id,
            )
        await db.commit()
        await db.refresh(reimb)

    return result_dict


@router.put("/{reimb_id}/invoices/{invoice_id}/voucher-review", response_model=VoucherReviewResponse)
async def review_reimbursement_invoice(
    reimb_id: int,
    invoice_id: int,
    body: VoucherReviewRequest,
    request: Request,
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """管理员在报销单随单审核中直接复核并代修正发票。"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="仅管理员可操作")

    reimb_query = select(Reimbursement).options(selectinload(Reimbursement.invoices)).where(
        Reimbursement.id == reimb_id
    )
    reimb_result = await db.execute(reimb_query)
    reimb = reimb_result.scalar_one_or_none()

    if not reimb:
        raise HTTPException(status_code=404, detail="报销单不存在")
    if reimb.status != ReimbursementStatus.SUBMITTED:
        raise HTTPException(status_code=400, detail="仅待审批报销单支持随单复核")

    invoice = next((inv for inv in reimb.invoices if inv.id == invoice_id), None)
    if not invoice:
        raise HTTPException(status_code=404, detail="该发票不属于当前报销单")

    diff_rows = (
        await db.execute(select(ParsingDiff).where(ParsingDiff.invoice_id == invoice_id))
    ).scalars().all()
    parsing_diffs = {diff.field_name: diff for diff in diff_rows}

    field_states = invoice.field_states or {}
    corrected_fields: List[str] = []
    admin_changes: List[dict] = []

    for update in body.field_updates:
        if update.source not in {"ocr", "llm", "custom"}:
            raise HTTPException(status_code=400, detail=f"字段 {update.field_name} 的来源无效")
        if update.source == "custom" and (update.value is None or str(update.value).strip() == ""):
            raise HTTPException(status_code=400, detail=f"字段 {update.field_name} 的自定义值不能为空")

        diff = parsing_diffs.get(update.field_name)
        if not diff:
            raise HTTPException(status_code=400, detail=f"字段 {update.field_name} 不在当前发票复核范围内")

        if update.source == "ocr":
            final_value = diff.ocr_value
        elif update.source == "llm":
            final_value = diff.llm_value
        else:
            final_value = update.value

        if final_value is None or str(final_value).strip() == "":
            raise HTTPException(status_code=400, detail=f"字段 {update.field_name} 的目标值为空，无法保存")

        old_value = _serialize_invoice_field(getattr(invoice, update.field_name, None))
        try:
            _set_invoice_field(invoice, update.field_name, str(final_value))
        except Exception as exc:
            raise HTTPException(status_code=400, detail=f"字段 {update.field_name} 的值格式不正确") from exc

        diff.final_value = str(final_value)
        diff.source = update.source
        diff.resolved = 1

        new_value = _serialize_invoice_field(getattr(invoice, update.field_name, None))
        if old_value != new_value:
            corrected_fields.append(update.field_name)

        state = field_states.get(update.field_name, {})
        state["reviewed_value"] = new_value
        state["reviewed_by"] = current_user.get("username")
        state["reviewed_at"] = datetime.now().isoformat()
        field_states[update.field_name] = state

        admin_changes.append(
            {
                "field_name": update.field_name,
                "label": state.get("label") or update.field_name,
                "old_value": old_value,
                "new_value": new_value,
                "source": update.source,
            }
        )

    trace = invoice.decision_trace or {}
    existing_changes = trace.get("admin_corrections") or []
    trace["admin_corrections"] = existing_changes + admin_changes
    trace["voucher_review"] = {
        "reviewed": body.mark_reviewed,
        "reviewed_by": current_user.get("username"),
        "reviewed_at": datetime.now().isoformat(),
        "review_note": body.review_note or "",
        "reviewed_fields": admin_changes,
    }
    trace["last_admin_action"] = "correct" if corrected_fields else "confirm"

    invoice.field_states = field_states
    invoice.decision_trace = trace
    if corrected_fields:
        invoice.confirmation_mode = "ADMIN_CORRECTION"

    _rebuild_invoice_hash(invoice)

    client_info = get_client_info(request)
    await log_audit_no_commit(
        db=db,
        entity_type="invoice",
        entity_id=invoice_id,
        action="voucher_review",
        new_value={
            "reimbursement_id": reimb_id,
            "reviewed": body.mark_reviewed,
            "corrected_fields": corrected_fields,
            "confirmation_mode": invoice.confirmation_mode or "USER_SELECTION",
            "review_note": body.review_note or "",
        },
        ip_address=client_info.get("ip_address"),
        user_agent=client_info.get("user_agent"),
    )

    await db.commit()

    return VoucherReviewResponse(
        reimbursement_id=reimb_id,
        invoice_id=invoice_id,
        reviewed=body.mark_reviewed,
        corrected_fields=corrected_fields,
        confirmation_mode=invoice.confirmation_mode or "USER_SELECTION",
        message="该票随单复核已保存" if not corrected_fields else "管理员修正已保存",
    )


# ============================================================
# 第四步（配套）：后端审批通过 / 驳回端点
# ============================================================

@router.put("/{reimb_id}/approve")
async def approve_reimbursement(
    reimb_id: int,
    body: dict,
    request: Request,
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """审批通过报销单"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="仅管理员可操作")
    query = select(Reimbursement).options(
        selectinload(Reimbursement.invoices)
    ).where(Reimbursement.id == reimb_id)
    result = await db.execute(query)
    reimb = result.scalar_one_or_none()

    if not reimb:
        raise HTTPException(status_code=404, detail="报销单不存在")
    if reimb.status != ReimbursementStatus.SUBMITTED:
        raise HTTPException(status_code=400, detail="只能审批状态为「待审批」的报销单")

    unreviewed_invoices = [
        inv.invoice_number or str(inv.id)
        for inv in reimb.invoices
        if _invoice_needs_voucher_review(inv) and not _invoice_reviewed(inv)
    ]
    if unreviewed_invoices:
        raise HTTPException(
            status_code=400,
            detail=f"仍有发票未完成随单复核：{', '.join(unreviewed_invoices)}",
        )

    reimb.status = ReimbursementStatus.APPROVED
    reimb.reviewer = body.get("reviewer", "系统")
    reimb.review_note = body.get("review_note", "")

    # 审批通过后，关联发票才正式变为「已报销」
    for inv in reimb.invoices:
        inv.status = InvoiceStatus.REIMBURSED

    # 预算/风险预警检查
    warnings = []
    if reimb.ai_risk_level and ("高" in str(reimb.ai_risk_level) or "危" in str(reimb.ai_risk_level)):
        warnings.append(f"⚠️ AI 风险评级：{reimb.ai_risk_level}")
    if reimb.project_code:
        project = await db.scalar(select(Project).where(Project.project_code == reimb.project_code))
        if project and project.budget > 0:
            used = await db.scalar(
                select(func.coalesce(func.sum(Reimbursement.total_amount), 0)).where(
                    Reimbursement.project_code == reimb.project_code,
                    Reimbursement.status.in_([ReimbursementStatus.SUBMITTED, ReimbursementStatus.APPROVED, ReimbursementStatus.COMPLETED]),
                )
            ) or 0
            if Decimal(str(used)) > project.budget:
                warnings.append(f"⚠️ 项目 {project.project_name} 预算已超支！预算 ¥{float(project.budget):.2f}，已使用 ¥{float(used):.2f}")

    client_info = get_client_info(request)
    await log_audit_no_commit(
        db=db,
        entity_type="reimbursement",
        entity_id=reimb_id,
        action="approve",
        new_value={
            "status": ReimbursementStatus.APPROVED.value,
            "review_note": reimb.review_note or "",
            "invoice_count": len(reimb.invoices),
        },
        ip_address=client_info.get("ip_address"),
        user_agent=client_info.get("user_agent"),
    )

    if reimb.submitter_id:
        await push_notification(
            db, reimb.submitter_id,
            title="报销单审批通过",
            message=f"您的报销单「{reimb.title}」已审批通过，金额 ¥{float(reimb.total_amount or 0):.2f}，预计 3 个工作日内到账。",
            entity_type="reimbursement", entity_id=reimb.id,
        )

    await db.commit()
    await db.refresh(reimb)

    return {
        "message": "审批通过",
        "reimbursement_id": reimb_id,
        "invoice_count": len(reimb.invoices),
        "warnings": warnings,
    }


@router.put("/{reimb_id}/reject")
async def reject_reimbursement(
    reimb_id: int,
    body: dict,
    request: Request,
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """驳回报销单，释放关联发票"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="仅管理员可操作")
    query = select(Reimbursement).options(
        selectinload(Reimbursement.invoices)
    ).where(Reimbursement.id == reimb_id)
    result = await db.execute(query)
    reimb = result.scalar_one_or_none()

    if not reimb:
        raise HTTPException(status_code=404, detail="报销单不存在")
    if reimb.status != ReimbursementStatus.SUBMITTED:
        raise HTTPException(status_code=400, detail="只能驳回状态为「待审批」的报销单")

    reimb.status = ReimbursementStatus.REJECTED
    reject_reason = body.get("reject_reason", "无")
    reimb.reject_reason = reject_reason

    for inv in reimb.invoices:
        inv.reimbursement_id = None
        inv.status = InvoiceStatus.CONFIRMED

    client_info = get_client_info(request)
    await log_audit_no_commit(
        db=db,
        entity_type="reimbursement",
        entity_id=reimb_id,
        action="reject",
        new_value={
            "status": ReimbursementStatus.REJECTED.value,
            "reject_reason": reject_reason,
            "invoice_count": len(reimb.invoices),
        },
        ip_address=client_info.get("ip_address"),
        user_agent=client_info.get("user_agent"),
    )

    if reimb.submitter_id:
        await push_notification(
            db, reimb.submitter_id,
            title="报销单已驳回",
            message=f"您的报销单「{reimb.title}」已被驳回。理由：{reject_reason}",
            entity_type="reimbursement", entity_id=reimb.id,
        )

    await db.commit()
    await db.refresh(reimb)

    return {
        "message": "已驳回，关联发票已释放",
        "reimbursement_id": reimb_id,
    }


@router.put("/{reimb_id}/complete")
async def complete_reimbursement(
    reimb_id: int,
    request: Request,
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """出纳确认打款，模拟银企直联转账，报销单进入已完成状态。"""
    from app.services.payment_service import mock_bank_transfer

    query = select(Reimbursement).options(
        selectinload(Reimbursement.invoices),
        selectinload(Reimbursement.bank_card),
    ).where(Reimbursement.id == reimb_id)
    result = await db.execute(query)
    reimb = result.scalar_one_or_none()

    if not reimb:
        raise HTTPException(status_code=404, detail="报销单不存在")
    if reimb.status != ReimbursementStatus.APPROVED:
        raise HTTPException(status_code=400, detail="只能对已通过的报销单操作")

    # 模拟银企直联打款
    payment_result = mock_bank_transfer(reimb, reimb.bank_card)

    reimb.status = ReimbursementStatus.COMPLETED
    reimb.payment_transaction_id = payment_result["transaction_id"]
    reimb.payment_time = payment_result["transfer_time"]
    reimb.payment_bank = payment_result["to_bank"]

    if reimb.submitter_id:
        pay_tail = payment_result['to_account'][-4:]
        pay_arrival = payment_result['estimated_arrival'][:16].replace('T', ' ')
        await push_notification(
            db, reimb.submitter_id,
            title="报销款已打款",
            message=f"您的报销单「{reimb.title}」¥{float(reimb.total_amount or 0):.2f} 已打入尾号 {pay_tail} 的{payment_result['to_bank']}，流水号 {payment_result['transaction_id']}，预计 {pay_arrival} 前到账。",
            entity_type="reimbursement", entity_id=reimb.id,
        )

    # 审计
    client_info = get_client_info(request)
    await log_audit_no_commit(
        db=db, entity_type="reimbursement", entity_id=reimb_id,
        action="complete",
        new_value={
            "status": ReimbursementStatus.COMPLETED.value,
            "transaction_id": payment_result["transaction_id"],
            "payment_bank": payment_result["to_bank"],
        },
        ip_address=client_info.get("ip_address"),
        user_agent=client_info.get("user_agent"),
    )

    # 自动冲销关联借款
    borrowing = None
    if reimb.borrowing_id:
        from app.models.borrowing import Borrowing, BorrowingStatus
        b_query = select(Borrowing).where(Borrowing.id == reimb.borrowing_id)
        b_result = await db.execute(b_query)
        borrowing = b_result.scalar_one_or_none()
        if borrowing:
            borrowing.status = BorrowingStatus.REPAID.value
            borrowing.reimbursement_id = reimb.id
            borrowing.repaid_amount = reimb.total_amount

            reimb_amt = float(reimb.total_amount)
            borrow_amt = float(borrowing.estimated_amount)

            if reimb_amt > borrow_amt:
                # 超额冲销：报销金额 > 借款金额，通知借款人和所有管理员
                excess = reimb_amt - borrow_amt
                await push_notification(
                    db, borrowing.user_id,
                    title="借款超额冲销提醒",
                    message=f"您的借款「{borrowing.title}」（借款金额 ¥{borrow_amt:.2f}）已被报销单 #{reimb.id} 超额冲销，冲销金额 ¥{reimb_amt:.2f}，超出 ¥{excess:.2f}，请关注。",
                    entity_type="borrowing", entity_id=borrowing.id,
                )
                await push_notification_to_admins(
                    db,
                    title="借款超额冲销预警",
                    message=f"员工 {reimb.submitter} 的借款「{borrowing.title}」（¥{borrow_amt:.2f}）被报销单 #{reimb.id} 超额冲销 ¥{reimb_amt:.2f}，超出 ¥{excess:.2f}",
                    entity_type="borrowing", entity_id=borrowing.id,
                )
            elif reimb_amt < borrow_amt:
                # 部分冲销：报销金额 < 借款金额，通知借款人还有未冲销余额
                shortfall = borrow_amt - reimb_amt
                await push_notification(
                    db, borrowing.user_id,
                    title="借款部分冲销提醒",
                    message=f"您的借款「{borrowing.title}」（借款金额 ¥{borrow_amt:.2f}）已被报销单 #{reimb.id} 部分冲销 ¥{reimb_amt:.2f}，尚有 ¥{shortfall:.2f} 未冲销，请关注。",
                    entity_type="borrowing", entity_id=borrowing.id,
                )
            else:
                # 等额冲销，正常通知
                await push_notification(
                    db, borrowing.user_id,
                    title="借款已冲销",
                    message=f"您的借款「{borrowing.title}」¥{borrow_amt:.2f} 已被报销单 #{reimb.id} 全额冲销，冲销金额 ¥{reimb_amt:.2f}",
                    entity_type="borrowing", entity_id=borrowing.id,
                )

    # 银行卡余额变动
    card = reimb.bank_card
    if not card and reimb.submitter_id:
        card = (await db.execute(
            select(BankCard).where(BankCard.user_id == reimb.submitter_id)
        )).scalars().first()

    if card:
        reimb_amt = float(reimb.total_amount or 0)
        # 如果有关联借款，实际需补打款 = 报销金额 - 借款金额（借款已提前拨付）
        # 如果无借款，直接打款报销全额
        if reimb.borrowing_id and borrowing:
            borrow_amt = float(borrowing.estimated_amount)
            borrow_repaid = float(borrowing.repaid_amount or 0)
            if reimb_amt > borrow_amt:
                # 报销超出借款，补差额
                diff = Decimal(str(reimb_amt - borrow_amt))
                balance_before = card.balance
                card.balance = Decimal(str(card.balance)) + diff
                db.add(Transaction(
                    type="报销到账",
                    amount=diff,
                    bank_card_id=card.id,
                    reimbursement_id=reimb.id,
                    borrowing_id=reimb.borrowing_id,
                    balance_before=balance_before,
                    balance_after=card.balance,
                    note=f"报销单「{reimb.title}」补差到账（报销{reimb_amt:.2f}，借款冲销{borrow_amt:.2f}）",
                ))
            # 借款冲销流水（不论是否超额，都要记录）
            db.add(Transaction(
                type="借款冲销",
                amount=Decimal(str(-borrow_repaid)),
                bank_card_id=card.id,
                reimbursement_id=reimb.id,
                borrowing_id=reimb.borrowing_id,
                balance_before=card.balance,
                balance_after=card.balance,
                note=f"借款「{borrowing.title}」被报销单「{reimb.title}」冲销 ¥{borrow_repaid:.2f}",
            ))
        else:
            # 无借款关联，直接打款
            balance_before = card.balance
            card.balance = Decimal(str(card.balance)) + Decimal(str(reimb.total_amount or 0))
            db.add(Transaction(
                type="报销到账",
                amount=reimb.total_amount,
                bank_card_id=card.id,
                reimbursement_id=reimb.id,
                balance_before=balance_before,
                balance_after=card.balance,
                note=f"报销单「{reimb.title}」全额到账",
            ))

    await db.commit()
    await db.refresh(reimb)

    return {"message": "已确认打款，报销单完成", "reimbursement_id": reimb_id, "payment": payment_result}


@router.get("/{reimb_id}/timeline")
async def get_reimbursement_timeline(
    reimb_id: int,
    db: AsyncSession = Depends(get_db),
):
    """获取报销单的资金追踪时间轴。"""
    # 1. 查报销单
    query = select(Reimbursement).where(Reimbursement.id == reimb_id)
    result = await db.execute(query)
    reimb = result.scalar_one_or_none()
    if not reimb:
        raise HTTPException(status_code=404, detail="报销单不存在")

    timeline = []

    # 2. 节点 1：提交报销
    timeline.append({
        "time": reimb.created_at.isoformat() if reimb.created_at else None,
        "status": "done",
        "title": "提交报销单",
        "description": f"报销事由：{reimb.title} | 金额：¥{float(reimb.total_amount or 0):.2f}",
    })

    # 3. 节点 2：AI 审查（如果有）
    if reimb.ai_risk_level:
        import json as _json
        risk_label = "低风险"
        status_icon = "done"
        if reimb.ai_risk_level and ("高" in str(reimb.ai_risk_level) or "危" in str(reimb.ai_risk_level)):
            risk_label = "高风险"
            status_icon = "error"

        # 从 JSON 中提取简洁摘要
        desc = "暂无审查意见"
        if reimb.ai_reason:
            try:
                ai_obj = _json.loads(reimb.ai_reason) if isinstance(reimb.ai_reason, str) else reimb.ai_reason
                parts = []
                if ai_obj.get("compliance_status"):
                    parts.append(ai_obj["compliance_status"])
                if ai_obj.get("reason"):
                    parts.append(ai_obj["reason"])
                desc = " · ".join(parts) if parts else str(reimb.ai_reason)[:120]
            except Exception:
                desc = str(reimb.ai_reason)[:120]

        timeline.append({
            "time": reimb.updated_at.isoformat() if reimb.updated_at else None,
            "status": status_icon,
            "title": f"AI 探针扫描 — {risk_label}",
            "description": desc,
        })

    # 4. 从审计日志中提取关键事件
    audit_query = select(AuditLog).where(
        AuditLog.entity_type == "reimbursement",
        AuditLog.entity_id == reimb_id,
    ).order_by(AuditLog.created_at.asc())
    audit_result = await db.execute(audit_query)
    audit_logs = audit_result.scalars().all()

    for log in audit_logs:
        if log.action == "approve":
            timeline.append({
                "time": log.created_at.isoformat() if log.created_at else None,
                "status": "done",
                "title": "财务审批通过",
                "description": log.details or "报销单已审批",
            })
        elif log.action == "reject":
            timeline.append({
                "time": log.created_at.isoformat() if log.created_at else None,
                "status": "error",
                "title": "审批驳回",
                "description": log.details or "报销单已被驳回",
            })

    # 5. 根据当前状态补充节点
    if reimb.status == ReimbursementStatus.SUBMITTED and not reimb.ai_risk_level:
        timeline.append({
            "time": None,
            "status": "processing",
            "title": "等待 AI 探针扫描",
            "description": "系统将自动进行合规审查",
        })
    elif reimb.status == ReimbursementStatus.SUBMITTED and reimb.ai_risk_level:
        timeline.append({
            "time": None,
            "status": "processing",
            "title": "等待财务总监审批",
            "description": f"审批人：{reimb.reviewer or 'admin（财务总监）'}",
        })
    elif reimb.status == ReimbursementStatus.APPROVED:
        timeline.append({
            "time": None,
            "status": "pending",
            "title": "预计打款到账",
            "description": "审批已通过，等待出纳打款",
        })
    elif reimb.status == ReimbursementStatus.COMPLETED:
        payment_desc = f"报销款 ¥{float(reimb.total_amount or 0):.2f} 已打入收款账户"
        if reimb.payment_transaction_id:
            payment_desc += f"，流水号 {reimb.payment_transaction_id}"
        if reimb.payment_bank:
            payment_desc += f"（{reimb.payment_bank}）"
        timeline.append({
            "time": reimb.payment_time.isoformat() if reimb.payment_time else (reimb.updated_at.isoformat() if reimb.updated_at else None),
            "status": "done",
            "title": "银企直联打款到账",
            "description": payment_desc,
        })
    elif reimb.status == ReimbursementStatus.REJECTED:
        timeline.append({
            "time": None,
            "status": "error",
            "title": "已被驳回",
            "description": reimb.reject_reason or "报销单被驳回，关联发票已释放",
        })

    return {"timeline": timeline, "reimbursement_id": reimb_id}


@router.get("/dashboard/stats")
async def get_dashboard_stats(db: AsyncSession = Depends(get_db)):
    """获取大屏的【全真】统计图表数据 (发票级细粒度拆分)"""
    # 已报销的发票总数：计数已通过/已打款报销单关联的发票
    reimbursed_invoice_count = await db.scalar(
        select(func.count(Invoice.id)).join(
            Reimbursement, Invoice.reimbursement_id == Reimbursement.id
        ).where(
            Reimbursement.status.in_([ReimbursementStatus.APPROVED, ReimbursementStatus.COMPLETED])
        )
    ) or 0
    total_reimbursed_amount = await db.scalar(
        select(func.coalesce(func.sum(Reimbursement.total_amount), 0)).where(
            Reimbursement.status.in_([ReimbursementStatus.APPROVED, ReimbursementStatus.COMPLETED])
        )
    ) or 0

    query = select(Reimbursement).options(
        selectinload(Reimbursement.invoices),
        selectinload(Reimbursement.submitter_user),
    )
    result = await db.execute(query)
    reimbs = result.scalars().all()

    trend_dict = {}
    pie_dict = {}
    ai_reject_count = 0

    for r in reimbs:
        # 月度趋势只统计已通过/已打款的报销，与 KPI 卡片口径一致
        if r.status not in (ReimbursementStatus.APPROVED, ReimbursementStatus.COMPLETED):
            continue
        dept = r.project_code or '通用部门'

        # ==========================================
        # 1. 柱状图真实数据：【细化到发票级别】进行打散统计
        # ==========================================
        if r.invoices:
            # 如果报销单底下有发票，就一张张遍历，分别计算月份和金额
            for inv in r.invoices:
                # 拿发票自己的时间，如果没有，降级用报销单的创建时间
                month = inv.issue_date.strftime('%Y-%m') if inv.issue_date else (
                    r.created_at.strftime('%Y-%m') if r.created_at else "2026-05")
                # 拿发票自己的金额 (确保转换成 float 进行累加)
                inv_amount = float(inv.total_with_tax or 0)

                trend_key = f"{month}_{dept}"
                if trend_key not in trend_dict:
                    trend_dict[trend_key] = {"month": month, "type": dept, "value": 0}

                # 精准累加这单张发票的金额
                trend_dict[trend_key]["value"] += inv_amount
        else:
            # 防御性编程：如果报销单刚建好，还没绑发票，就用报销单总金额和创建时间
            month = r.created_at.strftime('%Y-%m') if r.created_at else "2026-05"
            amount = float(r.total_amount or 0)
            trend_key = f"{month}_{dept}"
            if trend_key not in trend_dict:
                trend_dict[trend_key] = {"month": month, "type": dept, "value": 0}
            trend_dict[trend_key]["value"] += amount

        # ==========================================
        # 2. 饼图真实数据：按报销单的 AI 风险等级分组 (报销单级别)
        # ==========================================
        risk = getattr(r, 'ai_risk_level', None)
        if not risk:
            risk = "未经AI审查"

        if risk not in pie_dict:
            pie_dict[risk] = {"type": risk, "value": 0}

        # 饼图统计的是“单数”，所以每次 +1
        pie_dict[risk]["value"] += 1

        # 3. 真实统计高危拦截单量
        if risk in ["高", "高风险", "不合规", "高危"]:
            ai_reject_count += 1

    # 4. 预算使用率数据
    budget_data = []
    projects = (await db.execute(select(Project))).scalars().all()
    for p in projects:
        used = await db.scalar(
            select(func.coalesce(func.sum(Reimbursement.total_amount), 0)).where(
                Reimbursement.project_code == p.project_code,
                Reimbursement.status.in_([ReimbursementStatus.SUBMITTED, ReimbursementStatus.APPROVED, ReimbursementStatus.COMPLETED]),
            )
        ) or 0
        budget_data.append({
            "project_code": p.project_code,
            "project_name": p.project_name,
            "budget": float(p.budget),
            "used": float(used),
            "remaining": float(p.budget) - float(used),
            "usage_rate": round(float(used) / float(p.budget) * 100, 1) if float(p.budget) > 0 else 0,
        })

    # 5. 审批通过率（已通过+已打款 / 排除草稿和驳回的）
    total_reimbs = len(reimbs)
    approved_count = sum(1 for r in reimbs if r.status in [ReimbursementStatus.APPROVED, ReimbursementStatus.COMPLETED])
    decided_count = sum(1 for r in reimbs if r.status not in [ReimbursementStatus.DRAFT, ReimbursementStatus.SUBMITTED])
    approval_rate = round(approved_count / decided_count * 100, 1) if decided_count > 0 else 0

    # 6. 最近 10 条待审批
    pending_list = []
    for r in reimbs:
        if r.status == ReimbursementStatus.SUBMITTED:
            pending_list.append({
                "id": r.id,
                "title": r.title,
                "submitter": r.submitter_name or r.submitter or "-",
                "amount": float(r.total_amount or 0),
                "created_at": r.created_at.isoformat() if r.created_at else "",
            })
    pending_list.sort(key=lambda x: x["created_at"], reverse=True)
    pending_list = pending_list[:10]

    return {
        "trendData": list(trend_dict.values()),
        "pieData": list(pie_dict.values()),
        "aiRejectCount": ai_reject_count,
        "budgetData": budget_data,
        "approvalRate": approval_rate,
        "approvedCount": approved_count,
        "totalReimbCount": total_reimbs,
        "pendingList": pending_list,
        "reimbursedInvoiceCount": reimbursed_invoice_count,
        "totalReimbursedAmount": float(total_reimbursed_amount),
    }


@router.get("/dashboard/budget-prediction")
async def get_budget_prediction(db: AsyncSession = Depends(get_db)):
    """获取各项目预算耗尽预测（GM(1,1)+Markov 组合模型）"""
    from app.services.budget_prediction_service import predict_budget_exhaustion
    predictions = await predict_budget_exhaustion(db)
    return {"predictions": predictions}
