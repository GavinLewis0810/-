import hashlib
import json
from copy import deepcopy
from typing import Optional, List, Dict, Any, Tuple
from io import BytesIO
from datetime import date, datetime
from fastapi import APIRouter, Depends, UploadFile, File, HTTPException, Query, BackgroundTasks, Request
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, text
from sqlalchemy.orm.attributes import flag_modified
from decimal import Decimal

from app.database import get_db
from app.models.invoice import Invoice, OcrResult, LlmResult, ParsingDiff, InvoiceStatus
from app.models.image_forensics import ImageForensicsResult
from app.schemas.invoice import (
    InvoiceResponse, InvoiceListResponse, InvoiceDetailResponse,
    InvoiceUpdate, BatchUpdateRequest, BatchDeleteRequest, StatisticsResponse, UploadResponse,
    ResolveDiffRequest, GroundTruthSave,
    ConfirmInvoiceRequest, ConfirmInvoiceResponse,
    SubjectReviewApplyRequest, SubjectReviewApplyResponse,
)
from app.config import get_settings
from app.services.audit_service import log_audit_no_commit, get_client_info
from app.rate_limit import limiter
from app.dependencies import get_current_user

settings = get_settings()
router = APIRouter()


def _parse_invoice_ids(invoice_ids: Optional[str]) -> Optional[List[int]]:
    if not invoice_ids:
        return None
    ids: List[int] = []
    for raw_id in invoice_ids.split(","):
        raw_id = raw_id.strip()
        if not raw_id:
            continue
        try:
            ids.append(int(raw_id))
        except ValueError as exc:
            raise HTTPException(status_code=400, detail="发票ID必须为整数") from exc
    if not ids:
        raise HTTPException(status_code=400, detail="发票ID不能为空")
    return ids


def _parse_date_param(value: Optional[str], field_name: str) -> Optional[date]:
    if not value:
        return None
    try:
        return date.fromisoformat(value)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=f"{field_name} 日期格式无效，应为 YYYY-MM-DD") from exc


@router.post("/upload", response_model=List[UploadResponse])
@limiter.limit("10/minute")
async def upload_invoices(
        request: Request,
        background_tasks: BackgroundTasks,
        files: List[UploadFile] = File(...),
        db: AsyncSession = Depends(get_db),
        current_user: dict = Depends(get_current_user),
):
    """上传发票文件 (支持多文件)，上传后异步触发OCR解析"""
    results = []
    invoice_ids_to_process = []
    client_info = get_client_info(request)

    for file in files:
        # Validate file type
        ext = file.filename.split(".")[-1].lower() if file.filename else ""
        if ext not in settings.allowed_extensions:
            results.append(UploadResponse(
                id=0,
                file_name=file.filename or "unknown",
                status="error",
                message=f"不支持的文件类型: {ext}"
            ))
            continue

        # Read file content
        content = await file.read()

        # Validate file size
        if len(content) > settings.max_file_size:
            results.append(UploadResponse(
                id=0,
                file_name=file.filename or "unknown",
                status="error",
                message=f"文件过大，最大支持 {settings.max_file_size // 1024 // 1024}MB"
            ))
            continue

        # Create invoice record with UPLOADED status (not yet processed)
        invoice = Invoice(
            file_name=file.filename or "unknown",
            file_type=ext,
            file_data=content,
            status=InvoiceStatus.UPLOADED,
            owner=current_user["username"],     # 旧字段（迁移过渡期）
            owner_id=current_user["id"],        # 外键关联
        )
        db.add(invoice)
        await db.flush()

        invoice_ids_to_process.append(invoice.id)

        # Audit log for upload
        await log_audit_no_commit(
            db=db,
            entity_type="invoice",
            entity_id=invoice.id,
            action="upload",
            new_value={"file_name": invoice.file_name, "file_type": ext, "file_size": len(content)},
            ip_address=client_info.get("ip_address"),
            user_agent=client_info.get("user_agent"),
        )

        results.append(UploadResponse(
            id=invoice.id,
            file_name=invoice.file_name,
            status="success",
            message="上传成功，等待解析"
        ))

    await db.commit()

    # Schedule background processing for each uploaded invoice
    for invoice_id in invoice_ids_to_process:
        background_tasks.add_task(process_invoice_background, invoice_id)

    return results


async def process_invoice_background(invoice_id: int, max_retries: int = 3):
    """Background task to process an invoice with OCR/LLM."""
    from app.services.invoice_service import process_invoice as do_process
    from app.database import async_session_maker
    from app.services.audit_service import log_audit
    import logging
    import asyncio
    logger = logging.getLogger(__name__)

    retry_count = 0
    last_error = None

    while retry_count <= max_retries:
        async with async_session_maker() as db:
            try:
                # Update status to PROCESSING
                query = select(Invoice).where(Invoice.id == invoice_id)
                result = await db.execute(query)
                invoice = result.scalar_one_or_none()

                if not invoice:
                    logger.error(f"Invoice {invoice_id} not found")
                    return

                invoice.status = InvoiceStatus.PROCESSING
                await db.commit()

                logger.info(f"Background processing invoice {invoice_id} (attempt {retry_count + 1}/{max_retries + 1})")
                success = await do_process(invoice_id, db)

                if success:
                    logger.info(f"Invoice {invoice_id} processing completed successfully")
                    # Log successful processing
                    await log_audit(
                        db=db,
                        entity_type="invoice",
                        entity_id=invoice_id,
                        action="process_complete",
                        new_value={"status": "success", "attempts": retry_count + 1}
                    )
                    return
                else:
                    last_error = "Processing returned false"
                    logger.warning(f"Invoice {invoice_id} processing returned false")

            except Exception as e:
                last_error = str(e)
                logger.error(f"Failed to process invoice {invoice_id} (attempt {retry_count + 1}): {e}")
                await db.rollback()

        retry_count += 1

        if retry_count <= max_retries:
            # Exponential backoff: 2^retry_count seconds (2, 4, 8 seconds)
            delay = 2 ** retry_count
            logger.info(f"Retrying invoice {invoice_id} in {delay} seconds...")
            await asyncio.sleep(delay)

    # All retries exhausted - mark as failed
    async with async_session_maker() as db:
        try:
            query = select(Invoice).where(Invoice.id == invoice_id)
            result = await db.execute(query)
            invoice = result.scalar_one_or_none()

            if invoice:
                # Set status back to UPLOADED so user can retry manually
                invoice.status = InvoiceStatus.UPLOADED
                await db.commit()

                # Log failed processing
                await log_audit(
                    db=db,
                    entity_type="invoice",
                    entity_id=invoice_id,
                    action="process_failed",
                    new_value={"error": last_error, "attempts": max_retries + 1}
                )

            logger.error(f"Invoice {invoice_id} processing failed after {max_retries + 1} attempts: {last_error}")
        except Exception as e:
            logger.error(f"Failed to update invoice {invoice_id} status after retry exhaustion: {e}")


@router.get("", response_model=InvoiceListResponse)
async def list_invoices(
        page: int = Query(1, ge=1, description="页码"),
        page_size: int = Query(20, ge=1, le=100, description="每页数量"),
        status: Optional[InvoiceStatus] = Query(None, description="状态筛选"),
        owner: Optional[str] = Query(None, description="归属人筛选"),
        start_date: Optional[str] = Query(None, description="开始日期"),
        end_date: Optional[str] = Query(None, description="结束日期"),
        db: AsyncSession = Depends(get_db),
        current_user: dict = Depends(get_current_user),
):
    """获取发票列表。管理员看全部，员工只看自己的。"""
    query = select(Invoice)

    # 数据隔离：员工只能看归属于自己的发票（通过外键）
    if current_user["role"] != "admin":
        query = query.where(Invoice.owner_id == current_user["id"])

    # Apply filters
    if status:
        query = query.where(Invoice.status == status)
    if owner:
        query = query.where(Invoice.owner == owner)
    if start_date:
        try:
            start_date_obj = date.fromisoformat(start_date)
            query = query.where(Invoice.issue_date >= start_date_obj)
        except ValueError:
            pass  # Invalid date format, skip filter
    if end_date:
        try:
            end_date_obj = date.fromisoformat(end_date)
            query = query.where(Invoice.issue_date <= end_date_obj)
        except ValueError:
            pass  # Invalid date format, skip filter

    # Count total
    count_query = select(func.count()).select_from(query.subquery())
    total = await db.scalar(count_query) or 0

    # Apply pagination
    query = query.order_by(Invoice.created_at.desc())
    query = query.offset((page - 1) * page_size).limit(page_size)

    result = await db.execute(query)
    invoices = result.scalars().all()

    return InvoiceListResponse(
        items=[InvoiceResponse.model_validate(inv) for inv in invoices],
        total=total,
        page=page,
        page_size=page_size
    )


@router.get("/statistics", response_model=StatisticsResponse)
async def get_statistics(
        invoice_ids: Optional[str] = Query(None, description="发票ID列表，逗号分隔"),
        status: Optional[InvoiceStatus] = Query(None, description="状态筛选"),
        owner: Optional[str] = Query(None, description="归属人筛选"),
        db: AsyncSession = Depends(get_db)
):
    """获取发票统计数据"""
    query = select(Invoice)

    ids = _parse_invoice_ids(invoice_ids)
    if ids:
        query = query.where(Invoice.id.in_(ids))
    if status:
        query = query.where(Invoice.status == status)
    if owner:
        query = query.where(Invoice.owner == owner)

    result = await db.execute(query)
    invoices = result.scalars().all()

    count = len(invoices)
    total_amount = sum((inv.amount or Decimal(0)) for inv in invoices)
    total_tax = sum((inv.tax_amount or Decimal(0)) for inv in invoices)
    total_with_tax = sum((inv.total_with_tax or Decimal(0)) for inv in invoices)

    return StatisticsResponse(
        count=count,
        total_amount=total_amount,
        total_tax=total_tax,
        total_with_tax=total_with_tax
    )


@router.get("/{invoice_id}", response_model=InvoiceDetailResponse)
async def get_invoice(
        invoice_id: int,
        db: AsyncSession = Depends(get_db)
):
    """获取发票详情"""
    query = select(Invoice).where(Invoice.id == invoice_id)
    result = await db.execute(query)
    invoice = result.scalar_one_or_none()

    if not invoice:
        raise HTTPException(status_code=404, detail="发票不存在")

    # Load related data
    ocr_query = select(OcrResult).where(OcrResult.invoice_id == invoice_id)
    ocr_result = await db.execute(ocr_query)
    ocr = ocr_result.scalar_one_or_none()

    llm_query = select(LlmResult).where(LlmResult.invoice_id == invoice_id)
    llm_result = await db.execute(llm_query)
    llm = llm_result.scalar_one_or_none()

    diff_query = select(ParsingDiff).where(ParsingDiff.invoice_id == invoice_id)
    diff_result = await db.execute(diff_query)
    diffs = diff_result.scalars().all()

    forensics_query = select(ImageForensicsResult).where(ImageForensicsResult.invoice_id == invoice_id)
    forensics_result_data = await db.execute(forensics_query)
    forensics = forensics_result_data.scalar_one_or_none()

    # 🚨 终极修复：删除了原来那些散落的商品字段，加入了 items 数组
    invoice_dict = {
        "id": invoice.id,
        "file_name": invoice.file_name,
        "file_type": invoice.file_type,
        "status": invoice.status,
        "owner": invoice.owner,
        "invoice_number": invoice.invoice_number,
        "issue_date": invoice.issue_date,
        "buyer_name": invoice.buyer_name,
        "buyer_tax_id": invoice.buyer_tax_id,
        "seller_name": invoice.seller_name,
        "seller_tax_id": invoice.seller_tax_id,
        "total_with_tax": invoice.total_with_tax,
        "amount": invoice.amount,
        "tax_rate": invoice.tax_rate,
        "tax_amount": invoice.tax_amount,
        "items": invoice.items,  # 将 PostgreSQL 中的 JSONB 直接输出为数组
        "owner_id": invoice.owner_id,
        "reimbursement_id": invoice.reimbursement_id,
        "invoice_hash": invoice.invoice_hash,
        "ground_truth": invoice.ground_truth,
        "field_states": invoice.field_states,
        "user_corrections": invoice.user_corrections,
        "confirmation_mode": invoice.confirmation_mode,
        "decision_trace": invoice.decision_trace,
        "selection_fields": invoice.selection_fields,
        "created_at": invoice.created_at,
        "updated_at": invoice.updated_at,
        "ocr_result": ocr,
        "llm_result": llm,
        "parsing_diffs": list(diffs),
        "forensics_result": forensics,
    }

    return InvoiceDetailResponse.model_validate(invoice_dict)


@router.get("/{invoice_id}/file")
async def get_invoice_file(
        invoice_id: int,
        db: AsyncSession = Depends(get_db)
):
    """获取发票原始文件"""
    from fastapi.responses import Response
    from urllib.parse import quote

    query = select(Invoice).where(Invoice.id == invoice_id)
    result = await db.execute(query)
    invoice = result.scalar_one_or_none()

    if not invoice:
        raise HTTPException(status_code=404, detail="发票不存在")

    content_type_map = {
        "pdf": "application/pdf",
        "jpg": "image/jpeg",
        "jpeg": "image/jpeg",
        "png": "image/png"
    }

    # URL-encode filename for Content-Disposition header (RFC 5987)
    encoded_filename = quote(invoice.file_name)

    return Response(
        content=invoice.file_data,
        media_type=content_type_map.get(invoice.file_type, "application/octet-stream"),
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"}
    )


@router.get("/{invoice_id}/preview")
async def preview_invoice_file(
        invoice_id: int,
        db: AsyncSession = Depends(get_db)
):
    """内嵌预览用：返回 inline 票面内容。"""
    from fastapi.responses import Response
    from urllib.parse import quote

    query = select(Invoice).where(Invoice.id == invoice_id)
    result = await db.execute(query)
    invoice = result.scalar_one_or_none()

    if not invoice:
        raise HTTPException(status_code=404, detail="发票不存在")

    content_type_map = {
        "pdf": "application/pdf",
        "jpg": "image/jpeg",
        "jpeg": "image/jpeg",
        "png": "image/png"
    }

    encoded_filename = quote(invoice.file_name)

    return Response(
        content=invoice.file_data,
        media_type=content_type_map.get(invoice.file_type, "application/octet-stream"),
        headers={"Content-Disposition": f"inline; filename*=UTF-8''{encoded_filename}"}
    )


@router.put("/{invoice_id}", response_model=InvoiceResponse)
async def update_invoice(
        invoice_id: int,
        update_data: InvoiceUpdate,
        request: Request,
        db: AsyncSession = Depends(get_db)
):
    """更新发票信息"""
    query = select(Invoice).where(Invoice.id == invoice_id)
    result = await db.execute(query)
    invoice = result.scalar_one_or_none()

    if not invoice:
        raise HTTPException(status_code=404, detail="发票不存在")

    # 查重逻辑
    if update_data.invoice_number:
        duplicate_query = select(Invoice).where(
            Invoice.invoice_number == update_data.invoice_number,
            Invoice.id != invoice_id,
            Invoice.status.in_([InvoiceStatus.CONFIRMED, InvoiceStatus.REIMBURSED])
        )
        duplicate_result = await db.execute(duplicate_query)
        duplicate_invoice = duplicate_result.scalar_one_or_none()

        if duplicate_invoice:
            raise HTTPException(
                status_code=400,
                detail=f"保存失败！系统内已存在发票号为【{update_data.invoice_number}】的记录，请勿重复录入。"
            )

    # Capture old values for audit
    update_dict = update_data.model_dump(exclude_unset=True)
    old_values = {key: getattr(invoice, key) for key in update_dict.keys()}
    # Convert non-serializable types
    for key, value in old_values.items():
        if hasattr(value, 'value'):  # Enum
            old_values[key] = value.value
        elif hasattr(value, 'isoformat'):  # Date/DateTime
            old_values[key] = value.isoformat()
        elif isinstance(value, Decimal):
            old_values[key] = str(value)

    for key, value in update_dict.items():
        setattr(invoice, key, value)

    # Audit log
    client_info = get_client_info(request)
    new_values = update_dict.copy()
    for key, value in new_values.items():
        if hasattr(value, 'value'):  # Enum
            new_values[key] = value.value
        elif hasattr(value, 'isoformat'):  # Date/DateTime
            new_values[key] = value.isoformat()
        elif isinstance(value, Decimal):
            new_values[key] = str(value)

    await log_audit_no_commit(
        db=db,
        entity_type="invoice",
        entity_id=invoice_id,
        action="update",
        old_value=old_values,
        new_value=new_values,
        ip_address=client_info.get("ip_address"),
        user_agent=client_info.get("user_agent"),
    )

    await db.commit()
    await db.refresh(invoice)

    return InvoiceResponse.model_validate(invoice)


FIELD_LABELS = {
    "invoice_number": "发票号码", "issue_date": "开票日期", "buyer_name": "购买方",
    "buyer_tax_id": "购买方税号", "seller_name": "销售方", "seller_tax_id": "销售方税号",
    "total_with_tax": "价税合计", "amount": "不含税金额", "tax_rate": "税率", "tax_amount": "税额",
}

SUBJECT_FIELDS = {"buyer_name", "buyer_tax_id", "seller_name", "seller_tax_id"}

HIGH_CONFIDENCE = 0.95
LOW_CONFIDENCE = 0.70


def _build_field_states(diffs: list) -> dict:
    """从 ParsingDiff 列表构建字段级状态快照。"""
    states = {}
    diff_map = {d.field_name: d for d in diffs}

    for field, label in FIELD_LABELS.items():
        diff = diff_map.get(field)
        if not diff:
            continue

        ocr_val = diff.ocr_value
        llm_val = diff.llm_value
        conf = float(diff.confidence or 0)
        eng_match = (ocr_val == llm_val) if (ocr_val is not None and llm_val is not None) else None

        if conf >= HIGH_CONFIDENCE and eng_match is True:
            status = "locked"
        elif conf >= LOW_CONFIDENCE:
            status = "correctable"
        elif eng_match is False:
            status = "conflict"
        else:
            status = "correctable"

        states[field] = {
            "status": status,
            "label": label,
            "ocr": ocr_val,
            "llm": llm_val,
            "confidence": round(conf, 2),
        }

    return states


async def _supports_status_value(db: AsyncSession, status_value: str) -> bool:
    """Check whether current DB enum accepts a given invoice status label."""
    try:
        rows = await db.execute(text(
            """
            SELECT 1
            FROM pg_type t
            JOIN pg_enum e ON e.enumtypid = t.oid
            WHERE e.enumlabel = :label
            LIMIT 1
            """
        ), {"label": status_value})
        return rows.first() is not None
    except Exception:
        # If DB is not PostgreSQL or catalog query fails, stay conservative.
        return False


async def _ensure_pending_voucher_review_status(db: AsyncSession) -> bool:
    """Best-effort ensure DB enum for invoices.status contains '待随单审核'."""
    # SQLAlchemy Enum persists enum member name by default (e.g. PENDING_RECHECK).
    target_label = InvoiceStatus.PENDING_VOUCHER_REVIEW.name
    if await _supports_status_value(db, target_label):
        return True

    try:
        # Discover enum type name from invoices.status column, then add value.
        row = await db.execute(text(
            """
            SELECT t.typname
            FROM pg_attribute a
            JOIN pg_class c ON a.attrelid = c.oid
            JOIN pg_namespace n ON c.relnamespace = n.oid
            JOIN pg_type t ON a.atttypid = t.oid
            WHERE n.nspname = 'public'
              AND c.relname = 'invoices'
              AND a.attname = 'status'
            LIMIT 1
            """
        ))
        enum_name_row = row.first()
        if not enum_name_row:
            return False
        enum_name = enum_name_row[0]
        await db.execute(text(
            f"ALTER TYPE {enum_name} ADD VALUE IF NOT EXISTS '{target_label}'"
        ))
        return await _supports_status_value(db, target_label)
    except Exception:
        return False


@router.post("/{invoice_id}/confirm", response_model=ConfirmInvoiceResponse)
async def confirm_invoice(
        invoice_id: int,
        body: ConfirmInvoiceRequest,
        request: Request,
        db: AsyncSession = Depends(get_db)
):
    """用户确认发票数据。修正过的字段会导致发票进入『待重审』状态。"""
    query = select(Invoice).where(Invoice.id == invoice_id)
    result = await db.execute(query)
    invoice = result.scalar_one_or_none()
    if not invoice:
        raise HTTPException(status_code=404, detail="发票不存在")

    if invoice.status not in [InvoiceStatus.REVIEWING, InvoiceStatus.PENDING]:
        raise HTTPException(status_code=400, detail="当前发票状态不允许确认操作")
    old_status = invoice.status.value

    # 加载解析差异数据（用局部变量，避免给 ORM relationship 赋值触发懒加载炸 greenlet）
    diff_query = select(ParsingDiff).where(ParsingDiff.invoice_id == invoice_id)
    diff_result = await db.execute(diff_query)
    diffs = diff_result.scalars().all()

    # 1. 构建字段状态快照
    field_states = _build_field_states(diffs)
    invoice.field_states = field_states

    selection_sources = {"ocr", "llm", "custom"}
    selected_fields = sorted({
        d.field_name for d in diffs
        if d.resolved == 1
        and (d.source or "").lower() in selection_sources
        and (
            (d.machine_source or "").lower() in {"manual_review", "", "matched"}
            or (d.source or "").lower() != (d.machine_source or "").lower()
            or (d.final_value or None) != (d.machine_value or None)
        )
    })
    invoice.selection_fields = selected_fields or None

    # 2. 对比用户修正
    corrections = body.corrections or {}
    corrected_fields: list[str] = []
    for field, user_val in corrections.items():
        if field not in field_states:
            continue
        cur = getattr(invoice, field, None)
        cur_str = str(cur) if cur is not None else None
        if user_val != cur_str:
            corrected_fields.append(field)

    confirmation_mode = "AUTO"
    risk_level = "low"
    requires_voucher_review = False

    existing_trace = deepcopy(invoice.decision_trace) if isinstance(invoice.decision_trace, dict) else {}
    subject_review = existing_trace.get("subject_review") if existing_trace else None
    subject_selected = [field for field in selected_fields if field in SUBJECT_FIELDS]

    if corrected_fields:
        invoice.user_corrections = {f: corrections[f] for f in corrected_fields}
        for field in corrected_fields:
            setattr(invoice, field, corrections[field])
        invoice.status = InvoiceStatus.PENDING_RECHECK
        confirmation_mode = "USER_EDIT"
        risk_level = "high"
        message = f"已标记 {len(corrected_fields)} 个手工修正字段，进入待重审"
    elif selected_fields:
        invoice.user_corrections = None
        ensured = await _ensure_pending_voucher_review_status(db)
        if not ensured:
            raise HTTPException(status_code=500, detail="数据库状态枚举未完成升级，无法进入待随单审核")
        invoice.status = InvoiceStatus.PENDING_VOUCHER_REVIEW
        message = f"检测到用户选择字段 {len(selected_fields)} 个，进入待随单审核"
        confirmation_mode = "USER_SELECTION"
        risk_level = "high" if subject_selected else "medium"
        requires_voucher_review = True
    else:
        invoice.user_corrections = None
        invoice.status = InvoiceStatus.CONFIRMED
        confirmation_mode = "AUTO"
        risk_level = "low"
        message = "发票自动确认通过，可进入可报销池"

    invoice.confirmation_mode = confirmation_mode
    existing_trace = invoice.decision_trace if isinstance(invoice.decision_trace, dict) else {}
    invoice.decision_trace = {
        **existing_trace,
        "selected_fields": selected_fields,
        "corrected_fields": corrected_fields,
        "unresolved_diff_count": sum(1 for d in diffs if d.resolved == 0),
        "risk_level": risk_level,
        "requires_voucher_review": requires_voucher_review,
        "subject_review": {
            **subject_review,
            "user_selected_fields": subject_selected,
            "manual_confirmation_mode": confirmation_mode,
        } if isinstance(subject_review, dict) else None,
    }

    # 3. 审计日志
    client_info = get_client_info(request)
    await log_audit_no_commit(
        db=db, entity_type="invoice", entity_id=invoice_id,
        action="confirm",
        old_value={"status": old_status, "corrections": corrections},
        new_value={
            "status": invoice.status.value,
            "confirmation_mode": confirmation_mode,
            "corrected_fields": corrected_fields,
            "selection_fields": selected_fields,
            "risk_level": risk_level,
        },
        ip_address=client_info.get("ip_address"),
        user_agent=client_info.get("user_agent"),
    )

    await db.commit()
    await db.refresh(invoice)

    return ConfirmInvoiceResponse(
        invoice_id=invoice.id,
        status=invoice.status.value,
        has_corrections=len(corrected_fields) > 0,
        corrected_fields=corrected_fields,
        confirmation_mode=confirmation_mode,
        risk_level=risk_level,
        requires_voucher_review=requires_voucher_review,
        selection_fields=selected_fields,
        message=message,
        next_status_label=(
            "待重审" if corrected_fields
            else "待随单审核" if selected_fields
            else "已确认"
        ),
        workflow_transition=(
            "REVIEWING→PENDING_RECHECK" if corrected_fields
            else "REVIEWING→PENDING_VOUCHER_REVIEW" if selected_fields
            else "REVIEWING→CONFIRMED"
        ),
    )


@router.post("/{invoice_id}/ground-truth")
async def save_ground_truth(
        invoice_id: int,
        body: GroundTruthSave,
        request: Request,
        db: AsyncSession = Depends(get_db)
):
    """保存人工标注真值，用于双引擎精度评估"""
    query = select(Invoice).where(Invoice.id == invoice_id)
    result = await db.execute(query)
    invoice = result.scalar_one_or_none()

    if not invoice:
        raise HTTPException(status_code=404, detail="发票不存在")

    invoice.ground_truth = body.fields
    await db.commit()
    await db.refresh(invoice)

    client_info = get_client_info(request)
    await log_audit_no_commit(
        db=db, entity_type="invoice", entity_id=invoice_id,
        action="set_ground_truth",
        new_value=body.fields,
        ip_address=client_info.get("ip_address"),
        user_agent=client_info.get("user_agent"),
    )
    await db.commit()
    return {"message": "真值已保存", "invoice_id": invoice_id, "fields_count": len(body.fields)}


# ── 双引擎精度评估 ──────────────────────────────────────────────

_EVAL_FIELDS = [
    'invoice_number', 'issue_date', 'buyer_name', 'buyer_tax_id',
    'seller_name', 'seller_tax_id', 'total_with_tax', 'amount', 'tax_amount',
]
_EVAL_NUMERIC = {'total_with_tax', 'amount', 'tax_amount'}
_EVAL_LABELS = {
    'invoice_number': '发票号码', 'issue_date': '开票日期',
    'buyer_name': '购买方名称', 'buyer_tax_id': '购买方纳税人识别号',
    'seller_name': '销售方名称', 'seller_tax_id': '销售方纳税人识别号',
    'total_with_tax': '价税合计', 'amount': '总金额', 'tax_amount': '总税额',
}


def _build_engine_fields(row: Any, comparable_fields: List[str]) -> Dict[str, Any]:
    return {field: getattr(row, field, None) if row is not None else None for field in comparable_fields}


async def _collect_invoice_eval_context(db: AsyncSession, invoice: Invoice) -> Tuple[Any, Any, Dict[str, ParsingDiff]]:
    inv_id = invoice.id
    ocr_row = (await db.execute(select(OcrResult).where(OcrResult.invoice_id == inv_id))).scalar_one_or_none()
    llm_row = (await db.execute(select(LlmResult).where(LlmResult.invoice_id == inv_id))).scalar_one_or_none()
    diffs = (await db.execute(select(ParsingDiff).where(ParsingDiff.invoice_id == inv_id))).scalars().all()
    return ocr_row, llm_row, {diff.field_name: diff for diff in diffs}


def _extract_eval_values(
    invoice: Invoice,
    field: str,
    ocr_row: Any,
    llm_row: Any,
    diff_map: Dict[str, ParsingDiff],
) -> Dict[str, str]:
    diff = diff_map.get(field)
    return {
        'ocr': _norm(getattr(ocr_row, field, None) if ocr_row else None, field),
        'llm': _norm(getattr(llm_row, field, None) if llm_row else None, field),
        'machine': _norm(diff.machine_value if diff and diff.machine_value is not None else None, field),
        'final': _norm(
            diff.final_value if diff and diff.final_value is not None else getattr(invoice, field, None),
            field,
        ),
    }


def _rerun_machine_decisions(ocr_row: Any, llm_row: Any, diff_map: Dict[str, ParsingDiff]) -> Dict[str, Dict[str, Any]]:
    from app.services.invoice_service import COMPARABLE_FIELDS as SERVICE_FIELDS, _compare_and_resolve

    ocr_fields = _build_engine_fields(ocr_row, SERVICE_FIELDS)
    llm_fields = _build_engine_fields(llm_row, SERVICE_FIELDS)
    has_llm = llm_row is not None
    if not has_llm:
        return {
            field: {
                'field_name': field,
                'machine_value': ocr_fields.get(field),
                'machine_source': 'ocr',
                'machine_confidence': None,
                'decision_rule_type': 'single_engine',
                'decision_reason': ['llm_unavailable'],
            }
            for field in SERVICE_FIELDS
            if ocr_fields.get(field)
        }
    ocr_confs = {
        field: (float(diff_map[field].ocr_confidence) * 100.0) if field in diff_map and diff_map[field].ocr_confidence is not None else None
        for field in SERVICE_FIELDS
    }
    llm_confs = {
        field: float(diff_map[field].llm_confidence) if field in diff_map and diff_map[field].llm_confidence is not None else None
        for field in SERVICE_FIELDS
    }
    _, rerun_diffs, _ = _compare_and_resolve(ocr_fields, llm_fields, has_llm, llm_confs, ocr_confs)
    return {item['field_name']: item for item in rerun_diffs}


async def _build_fusion_experiment(db: AsyncSession) -> Dict[str, Any]:
    invoices = (await db.execute(select(Invoice).where(Invoice.ground_truth.isnot(None)))).scalars().all()
    if not invoices:
        return {"annotated_count": 0, "message": "暂无标注数据"}

    per_field = {field: {'ocr': 0, 'llm': 0, 'fusion': 0, 'total': 0} for field in _EVAL_FIELDS}
    overall = {'ocr': 0, 'llm': 0, 'fusion': 0, 'total': 0}
    typical_cases: List[Dict[str, Any]] = []

    for invoice in invoices:
        gt = invoice.ground_truth if isinstance(invoice.ground_truth, dict) else {}
        ocr_row, llm_row, diff_map = await _collect_invoice_eval_context(db, invoice)
        rerun_map = _rerun_machine_decisions(ocr_row, llm_row, diff_map)

        for field in _EVAL_FIELDS:
            gt_val = _norm(gt.get(field), field)
            if not gt_val:
                continue

            values = _extract_eval_values(invoice, field, ocr_row, llm_row, diff_map)
            fusion_decision = rerun_map.get(field, {})
            fusion_val = _norm(fusion_decision.get('machine_value'), field)

            ocr_ok = _eval_eq(values['ocr'], gt_val, field)
            llm_ok = _eval_eq(values['llm'], gt_val, field)
            fusion_ok = _eval_eq(fusion_val, gt_val, field)

            per_field[field]['ocr'] += int(ocr_ok)
            per_field[field]['llm'] += int(llm_ok)
            per_field[field]['fusion'] += int(fusion_ok)
            per_field[field]['total'] += 1

            overall['ocr'] += int(ocr_ok)
            overall['llm'] += int(llm_ok)
            overall['fusion'] += int(fusion_ok)
            overall['total'] += 1

            if (
                len(typical_cases) < 3
                and values['ocr']
                and values['llm']
                and not _eval_eq(values['ocr'], values['llm'], field)
                and fusion_decision.get('machine_source') in ('ocr', 'llm')
            ):
                typical_cases.append({
                    'invoice_id': invoice.id,
                    'file_name': invoice.file_name,
                    'field': field,
                    'label': _EVAL_LABELS.get(field, field),
                    'ocr_value': values['ocr'],
                    'llm_value': values['llm'],
                    'fusion_value': fusion_val,
                    'fusion_source': fusion_decision.get('machine_source'),
                    'ground_truth': gt_val,
                    'decision_rule_type': fusion_decision.get('decision_rule_type'),
                    'decision_reason': fusion_decision.get('decision_reason') or [],
                })

    total = overall['total']
    ocr_accuracy = round(overall['ocr'] / total, 4) if total else 0
    llm_accuracy = round(overall['llm'] / total, 4) if total else 0
    fusion_accuracy = round(overall['fusion'] / total, 4) if total else 0
    best_single = max(ocr_accuracy, llm_accuracy)

    field_rows = []
    for field in _EVAL_FIELDS:
        stats = per_field[field]
        if stats['total'] == 0:
            continue
        ocr_rate = round(stats['ocr'] / stats['total'], 4)
        llm_rate = round(stats['llm'] / stats['total'], 4)
        fusion_rate = round(stats['fusion'] / stats['total'], 4)
        best_field_single = max(ocr_rate, llm_rate)
        field_rows.append({
            'field': field,
            'label': _EVAL_LABELS.get(field, field),
            'ocr': ocr_rate,
            'llm': llm_rate,
            'fusion': fusion_rate,
            'gain': round(fusion_rate - best_field_single, 4),
            'samples': stats['total'],
        })

    return {
        'annotated_count': len(invoices),
        'total_fields': total,
        'overall': {
            'ocr': ocr_accuracy,
            'llm': llm_accuracy,
            'fusion': fusion_accuracy,
            'best_single': best_single,
            'fusion_gain': round(fusion_accuracy - best_single, 4),
        },
        'per_field': field_rows,
        'strategy_cards': [
            {'key': 'confidence', 'title': '置信度融合', 'desc': '综合 OCR 字段置信度与 LLM 自评置信度，形成候选基础分。'},
            {'key': 'validity', 'title': '字段合法性校验', 'desc': '针对税号、日期、票号、金额分别做格式与规则校验。'},
            {'key': 'consistency', 'title': '跨字段一致性', 'desc': '金额类字段校验 amount + tax_amount ≈ total_with_tax。'},
            {'key': 'risk', 'title': '高风险转人工', 'desc': '分数接近或规则无法判定时，不盲选，进入人工复核。'},
        ],
        'typical_cases': typical_cases,
    }


async def _build_workflow_metrics(db: AsyncSession) -> Dict[str, Any]:
    invoices = (await db.execute(select(Invoice))).scalars().all()
    if not invoices:
        return {"invoice_count": 0, "message": "暂无发票数据"}

    fields_total = 0
    conflict_count = 0
    auto_pass_count = 0
    manual_review_count = 0
    machine_conflict_decisions = 0
    machine_conflict_hits = 0
    annotated_final_total = 0
    annotated_final_correct = 0

    for invoice in invoices:
        gt = invoice.ground_truth if isinstance(invoice.ground_truth, dict) else {}
        ocr_row, llm_row, diff_map = await _collect_invoice_eval_context(db, invoice)
        rerun_map = _rerun_machine_decisions(ocr_row, llm_row, diff_map)

        for field, diff in diff_map.items():
            if field not in _EVAL_FIELDS:
                continue
            values = _extract_eval_values(invoice, field, ocr_row, llm_row, diff_map)
            machine_decision = rerun_map.get(field, {})
            has_candidate = bool(values['ocr'] or values['llm'])
            if not has_candidate:
                continue
            fields_total += 1

            is_conflict = bool(values['ocr'] and values['llm'] and not _eval_eq(values['ocr'], values['llm'], field))
            if is_conflict:
                conflict_count += 1

            if machine_decision.get('machine_source') == 'manual_review':
                manual_review_count += 1
            elif machine_decision.get('machine_source') in ('ocr', 'llm', 'matched') and machine_decision.get('machine_value'):
                auto_pass_count += 1

            gt_val = _norm(gt.get(field), field)
            if gt_val:
                annotated_final_total += 1
                if _eval_eq(values['final'], gt_val, field):
                    annotated_final_correct += 1

                if is_conflict and machine_decision.get('machine_source') in ('ocr', 'llm') and machine_decision.get('machine_value'):
                    machine_conflict_decisions += 1
                    if _eval_eq(_norm(machine_decision.get('machine_value'), field), gt_val, field):
                        machine_conflict_hits += 1

    return {
        'invoice_count': len(invoices),
        'fields_total': fields_total,
        'conflict_rate': round(conflict_count / fields_total, 4) if fields_total else 0,
        'auto_pass_rate': round(auto_pass_count / fields_total, 4) if fields_total else 0,
        'manual_review_rate': round(manual_review_count / fields_total, 4) if fields_total else 0,
        'auto_decision_hit_rate': round(machine_conflict_hits / machine_conflict_decisions, 4) if machine_conflict_decisions else 0,
        'final_human_in_loop_accuracy': round(annotated_final_correct / annotated_final_total, 4) if annotated_final_total else 0,
        'counts': {
            'conflict_count': conflict_count,
            'auto_pass_count': auto_pass_count,
            'manual_review_count': manual_review_count,
            'machine_conflict_decisions': machine_conflict_decisions,
        },
    }


@router.get("/eval/fusion-experiment")
async def eval_fusion_experiment(db: AsyncSession = Depends(get_db)):
    """离线纯算法实验：OCR vs LLM vs 融合策略。"""
    return await _build_fusion_experiment(db)


@router.get("/eval/workflow")
async def eval_workflow(db: AsyncSession = Depends(get_db)):
    """在线人机协同流程评估。"""
    return await _build_workflow_metrics(db)


@router.get("/eval/accuracy")
async def eval_accuracy(db: AsyncSession = Depends(get_db)):
    """兼容旧页面：同时返回离线实验与在线流程评估。"""
    return {
        'experiment': await _build_fusion_experiment(db),
        'workflow': await _build_workflow_metrics(db),
    }


def _norm(value, field: str) -> str:
    if value is None:
        return ""
    s = str(value).strip()
    if field in _EVAL_NUMERIC:
        s = s.replace('¥', '').replace('￥', '').replace(',', '').replace(' ', '')
        try:
            s = f"{float(s):.2f}"
        except ValueError:
            pass
    if field == 'issue_date':
        s = s.replace('/', '-').replace('.', '-')
    return s


def _eval_eq(v1: str, v2: str, field: str) -> bool:
    if not v1 and not v2:
        return True
    if not v1 or not v2:
        return False
    if field in _EVAL_NUMERIC:
        try:
            n1, n2 = float(v1), float(v2)
            if n2 == 0:
                return n1 == 0
            return abs(n1 - n2) / abs(n2) < 0.01
        except ValueError:
            return v1 == v2
    return v1 == v2


@router.post("/batch-update")
@limiter.limit("30/minute")
async def batch_update_invoices(
        request: Request,
        batch_request: BatchUpdateRequest,
        db: AsyncSession = Depends(get_db)
):
    """批量更新发票状态/归属人"""
    query = select(Invoice).where(Invoice.id.in_(batch_request.invoice_ids))
    result = await db.execute(query)
    invoices = result.scalars().all()

    client_info = get_client_info(request)
    updated_count = 0
    for invoice in invoices:
        old_values = {}
        new_values = {}
        if batch_request.status is not None:
            old_values["status"] = invoice.status.value if invoice.status else None
            invoice.status = batch_request.status
            new_values["status"] = batch_request.status.value
        if batch_request.owner is not None:
            old_values["owner"] = invoice.owner
            invoice.owner = batch_request.owner
            new_values["owner"] = batch_request.owner
        updated_count += 1

        # Audit log for each invoice
        await log_audit_no_commit(
            db=db,
            entity_type="invoice",
            entity_id=invoice.id,
            action="batch_update",
            old_value=old_values,
            new_value=new_values,
            ip_address=client_info.get("ip_address"),
            user_agent=client_info.get("user_agent"),
        )

    await db.commit()

    return {
        "message": f"成功更新 {updated_count} 张发票",
        "updated_count": updated_count
    }


@router.post("/batch-delete")
@limiter.limit("20/minute")
async def batch_delete_invoices(
        request: Request,
        batch_request: BatchDeleteRequest,
        db: AsyncSession = Depends(get_db)
):
    """批量删除发票及其关联数据"""
    if not batch_request.invoice_ids:
        raise HTTPException(status_code=400, detail="请选择要删除的发票")

    # Query invoices to delete
    query = select(Invoice).where(Invoice.id.in_(batch_request.invoice_ids))
    result = await db.execute(query)
    invoices = result.scalars().all()

    if not invoices:
        raise HTTPException(status_code=404, detail="未找到要删除的发票")

    client_info = get_client_info(request)
    deleted_count = 0
    for invoice in invoices:
        if invoice.reimbursement_id is not None:
            raise HTTPException(status_code=400, detail=f"发票 {invoice.invoice_number or invoice.id} 已被关联到报销单，无法删除。请先删除对应报销单")
        # Audit log for each deletion
        await log_audit_no_commit(
            db=db,
            entity_type="invoice",
            entity_id=invoice.id,
            action="delete",
            old_value={"file_name": invoice.file_name, "invoice_number": invoice.invoice_number},
            ip_address=client_info.get("ip_address"),
            user_agent=client_info.get("user_agent"),
        )
        await db.delete(invoice)  # cascade handles related records
        deleted_count += 1

    await db.commit()

    return {
        "message": f"成功删除 {deleted_count} 张发票",
        "deleted_count": deleted_count
    }


@router.post("/batch-reprocess")
@limiter.limit("5/minute")
async def batch_reprocess_invoices(
        request: Request,
        background_tasks: BackgroundTasks,
        batch_request: BatchDeleteRequest,  # Reuse for invoice_ids
        db: AsyncSession = Depends(get_db)
):
    """批量重新解析发票（清除旧的OCR/LLM结果，重新处理）"""
    import logging
    logger = logging.getLogger(__name__)

    if not batch_request.invoice_ids:
        raise HTTPException(status_code=400, detail="请选择要重新解析的发票")

    # Query invoices to reprocess
    query = select(Invoice).where(Invoice.id.in_(batch_request.invoice_ids))
    result = await db.execute(query)
    invoices = result.scalars().all()

    if not invoices:
        raise HTTPException(status_code=404, detail="未找到要重新解析的发票")

    # Clear old parsing results and reset invoice fields
    for invoice in invoices:
        # Delete old OCR results
        ocr_query = select(OcrResult).where(OcrResult.invoice_id == invoice.id)
        ocr_result = await db.execute(ocr_query)
        for ocr in ocr_result.scalars().all():
            await db.delete(ocr)

        # Delete old LLM results
        llm_query = select(LlmResult).where(LlmResult.invoice_id == invoice.id)
        llm_result = await db.execute(llm_query)
        for llm in llm_result.scalars().all():
            await db.delete(llm)

        # Delete old parsing diffs
        diff_query = select(ParsingDiff).where(ParsingDiff.invoice_id == invoice.id)
        diff_result = await db.execute(diff_query)
        for diff in diff_result.scalars().all():
            await db.delete(diff)

        # 🚨 修复重新解析时，清空新的字段
        invoice.invoice_number = None
        invoice.issue_date = None
        invoice.buyer_name = None
        invoice.buyer_tax_id = None
        invoice.seller_name = None
        invoice.seller_tax_id = None
        invoice.total_with_tax = None
        invoice.amount = None
        invoice.tax_amount = None
        invoice.tax_rate = None
        invoice.items = None  # 核心：清空旧的 JSONB 数组
        invoice.status = InvoiceStatus.UPLOADED

    await db.commit()
    logger.info(f"Cleared old parsing results for {len(invoices)} invoices, scheduling reprocess")

    # Schedule background processing
    for invoice in invoices:
        background_tasks.add_task(process_invoice_background, invoice.id)

    return {
        "message": f"已清除 {len(invoices)} 张发票的旧解析结果，正在重新解析",
        "count": len(invoices)
    }


@router.post("/{invoice_id}/process")
async def process_invoice(
        invoice_id: int,
        request: Request,
        db: AsyncSession = Depends(get_db)
):
    """处理发票：运行OCR解析"""
    from app.services.invoice_service import process_invoice as do_process

    query = select(Invoice).where(Invoice.id == invoice_id)
    result = await db.execute(query)
    invoice = result.scalar_one_or_none()

    if not invoice:
        raise HTTPException(status_code=404, detail="发票不存在")

    success = await do_process(invoice_id, db)

    if success:
        client_info = get_client_info(request)
        await log_audit_no_commit(
            db=db, entity_type="invoice", entity_id=invoice_id,
            action="process_complete",
            old_value={"status": "PROCESSING"},
            new_value={"status": "REVIEWING", "source": "ocr+llm"},
            ip_address=client_info.get("ip_address"),
            user_agent=client_info.get("user_agent"),
        )
        await db.commit()
        return {"message": "解析成功", "invoice_id": invoice_id}
    else:
        raise HTTPException(status_code=500, detail="解析失败")


@router.delete("/{invoice_id}")
async def delete_invoice(
        invoice_id: int,
        request: Request,
        db: AsyncSession = Depends(get_db)
):
    """删除发票"""
    query = select(Invoice).where(Invoice.id == invoice_id)
    result = await db.execute(query)
    invoice = result.scalar_one_or_none()

    if not invoice:
        raise HTTPException(status_code=404, detail="发票不存在")
    if invoice.reimbursement_id is not None:
        raise HTTPException(status_code=400, detail="该发票已被关联到报销单，无法删除。请先删除对应报销单")

    # Audit log
    client_info = get_client_info(request)
    await log_audit_no_commit(
        db=db,
        entity_type="invoice",
        entity_id=invoice_id,
        action="delete",
        old_value={"file_name": invoice.file_name, "invoice_number": invoice.invoice_number},
        ip_address=client_info.get("ip_address"),
        user_agent=client_info.get("user_agent"),
    )

    await db.delete(invoice)  # cascade deletes ai_call_logs, ocr/llm/diffs/forensics
    await db.commit()

    return {"message": "删除成功"}


@router.post("/{invoice_id}/diffs/{diff_id}/resolve")
async def resolve_diff(
        invoice_id: int,
        diff_id: int,
        resolve_request: ResolveDiffRequest,
        request: Request,
        db: AsyncSession = Depends(get_db)
):
    """解决解析差异，选择OCR、LLM或自定义值"""
    from datetime import datetime
    from decimal import Decimal as Dec

    # Get the diff
    diff_query = select(ParsingDiff).where(
        ParsingDiff.id == diff_id,
        ParsingDiff.invoice_id == invoice_id
    )
    diff_result = await db.execute(diff_query)
    diff = diff_result.scalar_one_or_none()

    if not diff:
        raise HTTPException(status_code=404, detail="差异记录不存在")

    # Capture old value for audit
    old_diff_value = {
        "field_name": diff.field_name,
        "final_value": diff.final_value,
        "source": diff.source,
        "resolved": diff.resolved,
    }

    # Determine the final value
    if resolve_request.source == 'ocr':
        final_value = diff.ocr_value
    elif resolve_request.source == 'llm':
        final_value = diff.llm_value
    elif resolve_request.source == 'custom':
        final_value = resolve_request.custom_value
    else:
        raise HTTPException(status_code=400, detail="无效的来源类型")

    # Update the diff
    diff.final_value = final_value
    diff.source = resolve_request.source
    diff.resolved = 1

    # Get the invoice and update the corresponding field
    invoice_query = select(Invoice).where(Invoice.id == invoice_id)
    invoice_result = await db.execute(invoice_query)
    invoice = invoice_result.scalar_one_or_none()

    if not invoice:
        raise HTTPException(status_code=404, detail="发票不存在")

    # Update the invoice field based on field_name
    field_name = diff.field_name
    if final_value is not None:
        if field_name == 'issue_date':
            try:
                invoice.issue_date = datetime.strptime(final_value, '%Y-%m-%d').date()
            except ValueError:
                pass
        # 🚨 修复：在比对逻辑中去掉了 quantity 和 unit_price
        elif field_name in ['total_with_tax', 'amount', 'tax_amount']:
            try:
                setattr(invoice, field_name, Dec(final_value))
            except (ValueError, TypeError):
                pass
        else:
            setattr(invoice, field_name, final_value)

    # Check if all diffs are resolved
    all_diffs_query = select(ParsingDiff).where(ParsingDiff.invoice_id == invoice_id)
    all_diffs_result = await db.execute(all_diffs_query)
    all_diffs = all_diffs_result.scalars().all()

    if isinstance(invoice.decision_trace, dict):
        existing_trace = deepcopy(invoice.decision_trace)
        subject_review = existing_trace.get("subject_review")
        if isinstance(subject_review, dict):
            subject_diffs = [item for item in all_diffs if item.field_name in SUBJECT_FIELDS]
            subject_review["resolved_fields"] = sorted([item.field_name for item in subject_diffs if item.resolved == 1])
            subject_review["all_fields_resolved"] = all(item.resolved == 1 for item in subject_diffs) if subject_diffs else False
            subject_review["last_action_source"] = resolve_request.source if field_name in SUBJECT_FIELDS else subject_review.get("last_action_source")
            invoice.decision_trace = {
                **existing_trace,
                "subject_review": subject_review,
            }
            flag_modified(invoice, "decision_trace")

    all_resolved = all(d.resolved == 1 for d in all_diffs)
    if all_resolved:
        invoice.status = InvoiceStatus.REVIEWING

    # Audit log for diff resolution
    client_info = get_client_info(request)
    await log_audit_no_commit(
        db=db,
        entity_type="parsing_diff",
        entity_id=diff_id,
        action="resolve",
        old_value=old_diff_value,
        new_value={
            "field_name": field_name,
            "final_value": final_value,
            "source": resolve_request.source,
            "resolved": 1,
            "invoice_id": invoice_id,
        },
        ip_address=client_info.get("ip_address"),
        user_agent=client_info.get("user_agent"),
    )

    await db.commit()

    return {
        "message": "差异已解决",
        "field_name": field_name,
        "final_value": final_value,
        "all_resolved": all_resolved
    }


@router.post("/{invoice_id}/subject-review/apply", response_model=SubjectReviewApplyResponse)
async def apply_subject_review(
    invoice_id: int,
    body: SubjectReviewApplyRequest,
    request: Request,
    db: AsyncSession = Depends(get_db),
):
    """主体复核整组应用：一次调用解决4个主体字段，避免前端循环4次逐字段 resolve。"""
    query = select(Invoice).where(Invoice.id == invoice_id)
    result = await db.execute(query)
    invoice = result.scalar_one_or_none()
    if not invoice:
        raise HTTPException(status_code=404, detail="发票不存在")

    existing_trace = deepcopy(invoice.decision_trace) if isinstance(invoice.decision_trace, dict) else {}
    subject_review = existing_trace.get("subject_review") if existing_trace else None
    if not subject_review:
        raise HTTPException(status_code=400, detail="该发票无主体复核数据")

    subject_diffs_result = await db.execute(
        select(ParsingDiff).where(
            ParsingDiff.invoice_id == invoice_id,
            ParsingDiff.field_name.in_(SUBJECT_FIELDS),
        )
    )
    subject_diffs = subject_diffs_result.scalars().all()
    diff_map = {d.field_name: d for d in subject_diffs}

    resolved_fields: List[str] = []
    scheme_key: Optional[str] = None
    scheme_display_label: Optional[str] = None

    if body.mode == "manual" and body.fields:
        # 手动修正模式
        for field_name in SUBJECT_FIELDS:
            value = body.fields.get(field_name)
            # 同步主表字段（含空值）
            setattr(invoice, field_name, value)
            # 标记 diff 为已确认
            diff = diff_map.get(field_name)
            if diff:
                diff.final_value = value
                diff.source = "custom"
                diff.resolved = 1
            resolved_fields.append(field_name)

        applied_mode = "manual"
        msg = "主体信息已手动修正"

    elif body.scheme_key:
        # 整组方案应用
        schemes = subject_review.get("candidate_schemes", [])
        scheme = next((s for s in schemes if s["key"] == body.scheme_key), None)
        if not scheme:
            raise HTTPException(status_code=400, detail=f"未知方案: {body.scheme_key}")

        scheme_key = scheme["key"]
        scheme_display_label = scheme.get("display_label", scheme["label"])

        for field_name in SUBJECT_FIELDS:
            value = scheme["fields"].get(field_name)
            # 关键修复：空值也必须同步主表和标记已确认，不能 continue 跳过
            setattr(invoice, field_name, value)
            diff = diff_map.get(field_name)
            if diff:
                diff.final_value = value
                diff.source = scheme["origins"].get(field_name, "ocr")
                diff.resolved = 1
            resolved_fields.append(field_name)

        applied_mode = "scheme"
        msg = "主体信息已确认"

    else:
        raise HTTPException(status_code=400, detail="请提供 scheme_key 或 mode=manual + fields")

    # 关键：主体复核完成 ≠ 整票确认完成，不在此处改 invoice.status
    # 状态变更留给 confirm_invoice 统一处理

    # 更新 subject_review 完整状态
    subject_review["applied"] = True
    subject_review["applied_mode"] = applied_mode
    subject_review["applied_scheme_key"] = scheme_key
    subject_review["applied_scheme_display_label"] = scheme_display_label
    subject_review["resolved_fields"] = resolved_fields
    subject_review["all_fields_resolved"] = True
    subject_review["manual_fields_changed"] = resolved_fields if body.mode == "manual" else []
    subject_review["applied_at"] = datetime.now().isoformat()
    subject_review["applied_operator"] = "admin"

    existing_trace["subject_review"] = subject_review
    invoice.decision_trace = existing_trace
    flag_modified(invoice, "decision_trace")

    # 审计日志
    client_info = get_client_info(request)
    await log_audit_no_commit(
        db=db,
        entity_type="invoice",
        entity_id=invoice_id,
        action="subject_review_apply",
        old_value={"status": invoice.status.value if invoice.status else None},
        new_value={
            "applied_mode": applied_mode,
            "applied_scheme_key": scheme_key,
            "scheme_display_label": scheme_display_label,
            "risk_level": subject_review.get("risk_level"),
            "manual_fields_changed": resolved_fields if body.mode == "manual" else [],
            "resolved_fields": resolved_fields,
            "next_status": invoice.status.value if invoice.status else "",
        },
        ip_address=client_info.get("ip_address"),
        user_agent=client_info.get("user_agent"),
        details=f"主体复核{'手动修正' if body.mode == 'manual' else '整组应用'}: {', '.join(resolved_fields)}",
    )

    await db.commit()

    return SubjectReviewApplyResponse(
        invoice_id=invoice_id,
        applied=True,
        applied_mode=applied_mode,
        scheme_key=scheme_key,
        scheme_display_label=scheme_display_label,
        resolved_fields=resolved_fields,
        all_subject_fields_resolved=True,
        next_status=invoice.status.value if invoice.status else "",
        next_status_label=invoice.status.value if invoice.status else "",
        message=msg,
    )


@router.post("/auto-confirm")
async def auto_confirm_invoices(
    request: Request,
    batch_request: BatchDeleteRequest,  # Reuse for invoice_ids
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """一键确认：自动确认所有无 OCR/LLM 冲突的「待确认」发票。"""
    invoice_ids = batch_request.invoice_ids
    if not invoice_ids:
        raise HTTPException(status_code=400, detail="请选择要确认的发票")

    query = select(Invoice).where(
        Invoice.id.in_(invoice_ids),
        Invoice.owner_id == current_user["id"],
        Invoice.status == InvoiceStatus.REVIEWING,
    )
    result = await db.execute(query)
    invoices = result.scalars().all()

    confirmed_ids = []
    need_manual_ids = []

    for invoice in invoices:
        # 检查是否有未解决的差异
        diff_query = select(ParsingDiff).where(
            ParsingDiff.invoice_id == invoice.id,
            ParsingDiff.resolved == 0,
        )
        diff_result = await db.execute(diff_query)
        unresolved_diffs = diff_result.scalars().all()

        # 检查必填字段
        critical_fields = [
            "invoice_number", "issue_date", "total_with_tax",
            "buyer_name", "seller_name",
        ]
        missing = [f for f in critical_fields if not getattr(invoice, f, None)]

        if unresolved_diffs or missing:
            need_manual_ids.append(invoice.id)
            continue

        # 无冲突、无缺失 → 自动确认
        raw = json.dumps({
            "invoice_number": invoice.invoice_number,
            "issue_date": str(invoice.issue_date) if invoice.issue_date else None,
            "total_with_tax": str(invoice.total_with_tax) if invoice.total_with_tax else None,
            "amount": str(invoice.amount) if invoice.amount else None,
            "tax_amount": str(invoice.tax_amount) if invoice.tax_amount else None,
            "buyer_name": invoice.buyer_name,
            "buyer_tax_id": invoice.buyer_tax_id,
            "seller_name": invoice.seller_name,
            "seller_tax_id": invoice.seller_tax_id,
            "items": json.dumps(invoice.items, ensure_ascii=False) if invoice.items else None,
            "owner_id": invoice.owner_id,
            "status": InvoiceStatus.CONFIRMED.value,
        }, sort_keys=True, ensure_ascii=False)
        invoice.invoice_hash = hashlib.sha256(raw.encode('utf-8')).hexdigest()

        invoice.status = InvoiceStatus.CONFIRMED
        invoice.confirmation_mode = "AUTO"
        invoice.selection_fields = None
        invoice.user_corrections = None
        existing_trace = invoice.decision_trace if isinstance(invoice.decision_trace, dict) else {}
        invoice.decision_trace = {
            **existing_trace,
            "selected_fields": [],
            "corrected_fields": [],
            "unresolved_diff_count": 0,
            "risk_level": "low",
            "requires_voucher_review": False,
            "source": "auto_confirm",
        }
        confirmed_ids.append(invoice.id)

        # 审计日志
        client_info = get_client_info(request)
        await log_audit_no_commit(
            db=db,
            entity_type="invoice",
            entity_id=invoice.id,
            action="auto_confirm",
            new_value={"status": InvoiceStatus.CONFIRMED.value},
            ip_address=client_info.get("ip_address"),
            user_agent=client_info.get("user_agent"),
        )

    await db.commit()

    return {
        "message": f"自动确认 {len(confirmed_ids)} 张发票",
        "confirmed_ids": confirmed_ids,
        "need_manual_ids": need_manual_ids,
    }


@router.get("/{invoice_id}/verify")
async def verify_invoice_integrity(
    invoice_id: int,
    db: AsyncSession = Depends(get_db),
):
    """验证发票数据完整性：重新计算哈希并与存证哈希对比。"""
    query = select(Invoice).where(Invoice.id == invoice_id)
    result = await db.execute(query)
    invoice = result.scalar_one_or_none()

    if not invoice:
        raise HTTPException(status_code=404, detail="发票不存在")
    if not invoice.invoice_hash:
        raise HTTPException(status_code=400, detail="该发票尚未生成数字指纹，请先确认发票")

    raw = json.dumps({
        "invoice_number": invoice.invoice_number,
        "issue_date": str(invoice.issue_date) if invoice.issue_date else None,
        "total_with_tax": str(invoice.total_with_tax) if invoice.total_with_tax else None,
        "amount": str(invoice.amount) if invoice.amount else None,
        "tax_amount": str(invoice.tax_amount) if invoice.tax_amount else None,
        "buyer_name": invoice.buyer_name,
        "buyer_tax_id": invoice.buyer_tax_id,
        "seller_name": invoice.seller_name,
        "seller_tax_id": invoice.seller_tax_id,
        "items": json.dumps(invoice.items, ensure_ascii=False) if invoice.items else None,
        "owner_id": invoice.owner_id,
        "status": invoice.status.value if invoice.status else None,
    }, sort_keys=True, ensure_ascii=False)
    current_hash = hashlib.sha256(raw.encode('utf-8')).hexdigest()

    is_valid = current_hash == invoice.invoice_hash

    return {
        "invoice_id": invoice_id,
        "valid": is_valid,
        "stored_hash": invoice.invoice_hash,
        "current_hash": current_hash,
        "message": "数字指纹校验通过，该票据存证后未被篡改。" if is_valid else "警告：数据已被篡改！当前哈希与存证哈希不一致。",
    }


@router.get("/export/csv")
@limiter.limit("10/minute")
async def export_invoices_csv(
        request: Request,
        invoice_ids: Optional[str] = Query(None, description="发票ID列表，逗号分隔"),
        status: Optional[InvoiceStatus] = Query(None, description="状态筛选"),
        owner: Optional[str] = Query(None, description="归属人筛选"),
        start_date: Optional[str] = Query(None, description="开始日期"),
        end_date: Optional[str] = Query(None, description="结束日期"),
        db: AsyncSession = Depends(get_db)
):
    """导出发票为CSV格式"""
    import csv
    from urllib.parse import quote

    query = select(Invoice)

    ids = _parse_invoice_ids(invoice_ids)
    if ids:
        query = query.where(Invoice.id.in_(ids))
    if status:
        query = query.where(Invoice.status == status)
    if owner:
        query = query.where(Invoice.owner == owner)
    start_date_obj = _parse_date_param(start_date, "start_date")
    if start_date_obj:
        query = query.where(Invoice.issue_date >= start_date_obj)
    end_date_obj = _parse_date_param(end_date, "end_date")
    if end_date_obj:
        query = query.where(Invoice.issue_date <= end_date_obj)

    query = query.order_by(Invoice.created_at.desc())
    result = await db.execute(query)
    invoices = result.scalars().all()

    output = BytesIO()
    output.write(b'\xef\xbb\xbf')

    import codecs
    writer = csv.writer(codecs.getwriter('utf-8')(output))

    writer.writerow([
        '发票号码', '开票日期', '购买方名称', '购买方纳税人识别号',
        '销售方名称', '销售方纳税人识别号', '包含商品明细',
        '金额', '税额', '价税合计', '税率',
        '状态', '归属人', '文件名', '创建时间'
    ])

    for inv in invoices:
        # 🚨 提取 JSONB 数组中所有的 item_name 并拼接成字符串
        item_names_str = ", ".join([str(i.get("item_name", "")) for i in inv.items if i.get("item_name")]) if inv.items else ""

        writer.writerow([
            inv.invoice_number or '',
            str(inv.issue_date) if inv.issue_date else '',
            inv.buyer_name or '',
            inv.buyer_tax_id or '',
            inv.seller_name or '',
            inv.seller_tax_id or '',
            item_names_str,  # 使用拼接后的名称
            str(inv.amount) if inv.amount else '',
            str(inv.tax_amount) if inv.tax_amount else '',
            str(inv.total_with_tax) if inv.total_with_tax else '',
            inv.tax_rate or '',
            inv.status.value if inv.status else '',
            inv.owner or '',
            inv.file_name or '',
            inv.created_at.strftime('%Y-%m-%d %H:%M:%S') if inv.created_at else ''
        ])

    output.seek(0)

    filename = quote('发票导出.csv')
    return StreamingResponse(
        output,
        media_type='text/csv; charset=utf-8',
        headers={
            'Content-Disposition': f"attachment; filename*=UTF-8''{filename}"
        }
    )


@router.get("/export/excel")
@limiter.limit("10/minute")
async def export_invoices_excel(
        request: Request,
        invoice_ids: Optional[str] = Query(None, description="发票ID列表，逗号分隔"),
        status: Optional[InvoiceStatus] = Query(None, description="状态筛选"),
        owner: Optional[str] = Query(None, description="归属人筛选"),
        start_date: Optional[str] = Query(None, description="开始日期"),
        end_date: Optional[str] = Query(None, description="结束日期"),
        db: AsyncSession = Depends(get_db)
):
    """导出发票为Excel格式"""
    from urllib.parse import quote

    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side
    except ImportError:
        raise HTTPException(status_code=500, detail="Excel导出需要安装openpyxl库")

    query = select(Invoice)

    ids = _parse_invoice_ids(invoice_ids)
    if ids:
        query = query.where(Invoice.id.in_(ids))
    if status:
        query = query.where(Invoice.status == status)
    if owner:
        query = query.where(Invoice.owner == owner)
    start_date_obj = _parse_date_param(start_date, "start_date")
    if start_date_obj:
        query = query.where(Invoice.issue_date >= start_date_obj)
    end_date_obj = _parse_date_param(end_date, "end_date")
    if end_date_obj:
        query = query.where(Invoice.issue_date <= end_date_obj)

    query = query.order_by(Invoice.created_at.desc())
    result = await db.execute(query)
    invoices = result.scalars().all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "发票列表"

    headers = [
        '发票号码', '开票日期', '购买方名称', '购买方纳税人识别号',
        '销售方名称', '销售方纳税人识别号', '包含商品明细',
        '金额', '税额', '价税合计', '税率',
        '状态', '归属人', '文件名', '创建时间'
    ]

    header_font = Font(bold=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

    for row, inv in enumerate(invoices, 2):
        # 🚨 提取 JSONB 数组中所有的 item_name 并拼接成字符串
        item_names_str = ", ".join([str(i.get("item_name", "")) for i in inv.items if i.get("item_name")]) if inv.items else ""

        data = [
            inv.invoice_number or '',
            str(inv.issue_date) if inv.issue_date else '',
            inv.buyer_name or '',
            inv.buyer_tax_id or '',
            inv.seller_name or '',
            inv.seller_tax_id or '',
            item_names_str, # 使用拼接后的名称
            float(inv.amount) if inv.amount else '',
            float(inv.tax_amount) if inv.tax_amount else '',
            float(inv.total_with_tax) if inv.total_with_tax else '',
            inv.tax_rate or '',
            inv.status.value if inv.status else '',
            inv.owner or '',
            inv.file_name or '',
            inv.created_at.strftime('%Y-%m-%d %H:%M:%S') if inv.created_at else ''
        ]

        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = thin_border

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = quote('发票导出.xlsx')
    return StreamingResponse(
        output,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={
            'Content-Disposition': f"attachment; filename*=UTF-8''{filename}"
        }
    )
